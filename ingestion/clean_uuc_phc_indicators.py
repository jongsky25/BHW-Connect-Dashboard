"""Apply the agreed indicator rules to GIDA reconciled data, and log every action.

Rules, set by the programme owner. The bound depends on what the indicator measures:

  Coverage percentages   Water, Pre-natal, SBA, FIC   capped at 100 - a share of a
                                                      population cannot exceed 100%.
  Rates per 1,000        IMR, UFMR, ABR               capped at 1,000 - a rate per
                                                      1,000 cannot exceed 1,000, but
                                                      may legitimately exceed 100.
  Any indicator          value below zero             set to 0.

Also applied:
  - The Pass/Fail and High/Low assessment columns are dropped. They were computed by
    the source office against the uncapped values and are not to be used.
  - "#N/A" reference values are written as blanks.
  - SORSOGON / PILAR / SAN ANTONIO resolves to PSGC 0506213048.
  - The provincial reference columns are collapsed to one row per province and written
    to a CSV. Water and Pre-natal references are capped at 100 (a provincial proportion
    cannot exceed 100% either); IMR, UFMR, ABR and FIC references are left as supplied.
  - Membership is aligned to the source office's final 2025 UUC for PHC list: five Cavite
    barangays it omits are dropped, and BASILAN / SUMISIP / SUMISIP CENTRAL, which it
    carries and the reconciled sheet does not, is recovered from the "2025 LIST" sheet.
    See FINAL_LIST_REMOVED / FINAL_LIST_ADDED below for what that costs.

No value is altered by the alignment and no indicator is dropped: the extract carries
5,987 rows, every one with its indicators populated.

Usage:
    python ingestion/clean_uuc_phc_indicators.py
"""
import argparse
import csv
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Column index in the "Reconciled" sheet -> label
CAP_100 = {33: 'Water', 27: 'Pre-natal', 30: 'SBA', 21: 'FIC'}
CAP_1000 = {15: 'IMR', 18: 'UFMR', 24: 'ABR'}
ALREADY_BOUNDED = {4: 'Physical Factor', 6: 'IP POP', 8: 'ARMED CONF', 9: 'IDP', 13: '4PS'}
INDICATORS = {**CAP_100, **CAP_1000, **ALREADY_BOUNDED}
CAPS = {**{i: 100 for i in CAP_100}, **{i: 1000 for i in CAP_1000}}

# Provincial reference columns. These are benchmarks, not barangay measurements, so they are
# NOT subject to the indicator rules — except Water and Pre-natal, which the owner capped at 100
# because a provincial proportion cannot exceed 100% either.
PROV_REF = {16: 'IMR', 19: 'UFMR', 22: 'FIC', 25: 'ABR',
            28: 'Pre-natal', 31: 'SBA', 34: 'Water'}
PROV_REF_CAP_100 = {28, 34}          # Pre-natal Prov Ref, Water Prov Ref

# Pass/Fail and High/Low assessment columns — dropped per owner decision
DROP_COLS = {5, 7, 11, 12, 14, 17, 20, 23, 26, 29, 32, 35, 38, 39, 40}

# One-off PSGC resolution supplied by the owner
PSGC_OVERRIDE = {('SORSOGON', 'PILAR', 'SAN ANTONIO'): '0506213048'}

# Machine-readable column names for the --data-csv extract (the loader's committed source).
# Explicit rather than slugified: the four geography columns carry the workbook's own pull-names,
# which are not usable as identifiers, and '4PS' cannot lead with a digit.
CSV_RENAME = {
    'Region': 'region_name',
    'province': 'province_name',
    'profile-brgy/muncity_pull': 'citymun_name',
    'profile-brgy/barangay_pull': 'barangay_name',
    '4PS': 'four_ps',
}

NEG_ACTION = 'Negative set to 0'

# ---------------------------------------------------------------- final-list alignment
#
# The reconciled workbook is the source for every *measurement* here, but it is not the source
# for who is ON the list. Its `NEW` sheet classifies 5,991 barangays UUA; the source office's
# final national list — "2025 UUC FOR PHC LIST", the one whose regional distribution the 2027
# Budget Cue Cards p37 publishes — carries 5,987, and the two differ on six barangays and
# nowhere else. This step reconciles the membership without touching a single value.
#
# Five barangays the reconciled workbook lists and the final list does not. All five are urban
# Cavite; all five read `UUA` in `NEW`. Dropped here, so they never reach the extract.
FINAL_LIST_REMOVED = {
    ('CAVITE', 'CITY OF BACOOR', 'MOLINO IV'),
    ('CAVITE', 'CITY OF BACOOR', 'SAN NICOLAS II'),
    ('CAVITE', 'CITY OF BACOOR', 'TALABA 2'),
    ('CAVITE', 'CITY OF BACOOR', 'TALABA 3'),
    ('CAVITE', 'CITY OF CAVITE', 'BARANGAY 38'),
}

# One barangay the final list carries and the reconciled workbook does not. `NEW` scores it
# `NOT UUA`, which is why it never entered this pipeline — but the same workbook's `2025 LIST`
# sheet holds a complete set of indicator values for it under PSGC 1900705019, and dim_geo
# already knows the code. So the row is recoverable rather than a hole, and it is recovered from
# there rather than invented.
#
# **Its values come from `2025 LIST`, not from the reconciled sheet, and that is a real
# difference in vintage.** The two disagree somewhere on 473 of the 5,989 barangays they share
# (ABR most often, 275 rows), so this one row is pre-reconciliation where the other 5,986 are
# post-. It is the only source that has it; the alternative is a listed barangay with no
# evidence at all. Recorded here, in the loader's docstring and in the migration that carries
# the change, so a reader of the row knows which extract it came from.
#
# Two of the workbook's columns have no counterpart in `2025 LIST` and are left empty rather
# than guessed:
#
#   ELCAC BRGY        — the conflict-area designation. Absent, so criterion (b) for this row
#                       rests on ARMED CONF + IDP alone (30 + 11 = 41, which passes anyway).
#   Health Indicators — the source office's own criterion (d) score, 0-7. It is a recorded
#                       classification that this pipeline never recomputes (see the column note
#                       in supabase/migrations/20260826150000_fact_uuc_phc_indicators.sql), and
#                       recomputing it here for one row is exactly what that note forbids. Left
#                       NULL, so route (d) does not count this barangay. Its listing does not
#                       depend on that: IP POP is 100, so criterion (a) carries it on its own.
#
# The seven provincial benchmarks are NOT taken from `2025 LIST`, which has none. They are
# province constants, so they are copied from the barangay's own province in this same extract —
# which is what keeps ref_uuc_phc_provincial's "no province carries two different reference
# values" assertion true for Basilan.
FINAL_LIST_ADDED = ('BASILAN', 'SUMISIP', 'SUMISIP CENTRAL')
FINAL_LIST_ADDED_PSGC = '1900705019'

# `2025 LIST` column index -> reconciled-sheet column index, for the values that carry over.
# FP CU (index 13 there) is deliberately absent: the reconciled extract drops it, so importing
# it would add a column the rest of the pipeline does not have.
ADDED_VALUE_COLS = {4: 4, 5: 6, 6: 8, 7: 9, 8: 13, 9: 15, 10: 18,
                    11: 21, 12: 24, 14: 27, 15: 30, 16: 33}


def align_to_final_list(rows, psgc_src):
    """Drop the five barangays the final list omits; add the one it adds. Mutates `rows`."""
    removed = [i for i, r in enumerate(rows)
               if (norm(r[1]), norm(r[2]), norm(r[3])) in FINAL_LIST_REMOVED]
    if len(removed) != len(FINAL_LIST_REMOVED):
        raise SystemExit(f'final list: expected {len(FINAL_LIST_REMOVED)} rows to remove, '
                         f'matched {len(removed)}')
    for i in reversed(removed):
        del rows[i]

    prov, citymun, brgy = FINAL_LIST_ADDED
    if any((norm(r[1]), norm(r[2]), norm(r[3])) == (prov, citymun, brgy) for r in rows):
        raise SystemExit(f'final list: {brgy} is already in the reconciled sheet')

    ws = openpyxl.load_workbook(psgc_src, read_only=True, data_only=True)['2025 LIST']
    src = next((r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[3] and str(r[3]).strip().zfill(10) == FINAL_LIST_ADDED_PSGC), None)
    if src is None:
        raise SystemExit(f'final list: {brgy} not found in the 2025 LIST sheet')

    sibling = next((r for r in rows if norm(r[1]) == prov), None)
    if sibling is None:
        raise SystemExit(f'final list: no {prov} row to copy provincial benchmarks from')

    new = [None] * len(sibling)
    new[0], new[1], new[2], new[3] = sibling[0], sibling[1], citymun, brgy
    for src_i, dst_i in ADDED_VALUE_COLS.items():
        new[dst_i] = src[src_i]
    for ci in PROV_REF:                       # province constants, copied not invented
        new[ci] = sibling[ci]
    new[41] = 'UUA'                           # DECISION — the final list is what says so

    # Keep the extract's own ordering: rows arrive grouped by municipality and sorted by
    # barangay within it, and every downstream diff reads better if that survives.
    group = [i for i, r in enumerate(rows)
             if (norm(r[1]), norm(r[2])) == (prov, citymun)]
    if not group:
        raise SystemExit(f'final list: no {prov} / {citymun} group to insert into')
    at = next((i for i in group if norm(rows[i][3]) > brgy), group[-1] + 1)
    rows.insert(at, new)
    return len(removed), 1



def norm(x):
    return str(x).strip().upper() if x is not None else ''


def csv_name(label):
    """Workbook header -> snake_case CSV column name."""
    if label in CSV_RENAME:
        return CSV_RENAME[label]
    return re.sub(r'_+', '_', re.sub(r'[^a-z0-9]+', '_', str(label).lower())).strip('_')


def is_na(v):
    return isinstance(v, str) and v.strip() == '#N/A'


def load_psgc(path):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)['2025 LIST']
    by3, by2 = {}, {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not (r[0] and str(r[0]).strip() not in ('PROVINCE', 'TOTAL') and r[1] and r[3]):
            continue
        code = str(r[3]).strip().zfill(10)
        by3[(norm(r[0]), norm(r[1]), norm(r[2]))] = code
        by2.setdefault((norm(r[1]), norm(r[2])), []).append(code)
    return by3, by2


def apply_rules(rows):
    """Mutate rows in place; return the actions taken."""
    actions = []
    for ri, r in enumerate(rows):
        for ci, label in INDICATORS.items():
            v = r[ci]
            if not isinstance(v, (int, float)):
                continue
            if v < 0:
                actions.append((ri, ci, label, v, NEG_ACTION, 0))
                r[ci] = 0
            elif ci in CAPS and v > CAPS[ci]:
                cap = CAPS[ci]
                actions.append((ri, ci, label, v, f'Capped at {cap:,}', cap))
                r[ci] = cap
        for ci in PROV_REF_CAP_100:         # provincial benchmarks that are proportions
            v = r[ci]
            if isinstance(v, (int, float)) and v > 100:
                r[ci] = 100
        for ci, v in enumerate(r):          # #N/A -> blank
            if is_na(v):
                r[ci] = None
    return actions


def provincial_reference(rows, HDR, before):
    """Collapse the denormalised Prov Ref columns into one row per province.

    `before` holds the pre-cap values so the sheet can report what was adjusted.
    Returns (records, n_capped_values).
    """
    seen, order = {}, []
    for ri, r in enumerate(rows):
        key = (str(r[0]).strip(), str(r[1]).strip() if r[1] else '')
        if key not in seen:
            seen[key] = {'region': key[0], 'province': key[1], 'barangays': 0,
                         'vals': {}, 'capped': set()}
            order.append(key)
        rec = seen[key]
        rec['barangays'] += 1
        for ci, lab in PROV_REF.items():
            if isinstance(r[ci], (int, float)) and lab not in rec['vals']:
                rec['vals'][lab] = r[ci]
                pre = before[ri][ci]
                if ci in PROV_REF_CAP_100 and isinstance(pre, (int, float)) and pre > 100:
                    rec['capped'].add(lab)
    recs = [seen[k] for k in sorted(order)]
    n_capped = sum(len(r['capped']) for r in recs)
    return recs, n_capped


# ---------------------------------------------------------------- styling
A = 'Arial'
H_FILL = PatternFill('solid', fgColor='1F3864')
H_FONT = Font(name=A, size=10, bold=True, color='FFFFFF')
CAP_FILL = PatternFill('solid', fgColor='FFF2CC')
TITLE = Font(name=A, size=14, bold=True, color='1F3864')
BOLD = Font(name=A, size=10, bold=True)
BODY = Font(name=A, size=10)
ITAL = Font(name=A, size=9, italic=True)
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def head(ws, labels, row=1, height=34):
    for j, h in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font, c.fill, c.border = H_FONT, H_FILL, BOX
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[row].height = height


def build(src, psgc_src, out_path, ref_csv=None, data_csv=None):
    rec = openpyxl.load_workbook(src, read_only=True, data_only=True)['Reconciled']
    allr = list(rec.iter_rows(min_row=1, values_only=True))
    HDR = list(allr[0])
    rows = [list(r) for r in allr[1:] if any(c is not None for c in r)]

    n_dropped, n_added = align_to_final_list(rows, psgc_src)

    by3, by2 = load_psgc(psgc_src)
    psgc_list, psgc_missing = [], 0
    for r in rows:
        k = (norm(r[1]), norm(r[2]), norm(r[3]))
        code = PSGC_OVERRIDE.get(k) or by3.get(k)
        if code is None:
            cands = by2.get((norm(r[2]), norm(r[3])), [])
            code = cands[0] if len(cands) == 1 else None
        if code is None:
            psgc_missing += 1
        psgc_list.append(code or '')

    n_na = sum(1 for r in rows for v in r if is_na(v))
    before = [list(r) for r in rows]                     # snapshot for the reference sheet
    actions = apply_rules(rows)
    prov_recs, n_prov_capped = provincial_reference(rows, HDR, before)
    n_cap100 = sum(1 for a in actions if a[1] in CAP_100 and a[4].startswith('Capped'))
    n_cap1000 = sum(1 for a in actions if a[1] in CAP_1000 and a[4].startswith('Capped'))
    n_neg = sum(1 for a in actions if a[4] == NEG_ACTION)
    touched = len({a[0] for a in actions})
    keep_idx = [i for i in range(len(HDR)) if i not in DROP_COLS]
    print(f'rows={len(rows)} final_list_dropped={n_dropped} final_list_added={n_added} '
          f'psgc_missing={psgc_missing} cap100={n_cap100} cap1000={n_cap1000} '
          f'neg={n_neg} na_blanked={n_na} cols_dropped={len(DROP_COLS)} touched={touched} '
          f'provinces={len(prov_recs)} prov_ref_capped={n_prov_capped}')

    out = openpyxl.Workbook()

    # ============================================================ README
    s = out.active
    s.title = 'README'
    s.sheet_view.showGridLines = False
    lines = [
        ('UUC for PHC 2025 — cleaned indicator dataset', 'title'),
        ('', None),
        ('Source', 'h'),
        ('GIDA reconciled data.xlsx, sheet "Reconciled" — 5,991 UUA barangays.', 'p'),
        ('PSGC codes joined from Submissions_UUA_2025_filled_1.xlsx, sheet "2025 LIST".', 'p'),
        ('Membership aligned to the final 2025 UUC for PHC list: 5,987 barangays.', 'p'),
        ('', None),
        ('Rules applied', 'h'),
        ('The bound depends on what the indicator measures.', 'p'),
        ('', None),
        ('  Water, Pre-natal, SBA, FIC   capped at 100. Coverage percentages — the share of', 'p'),
        ('                               households, mothers or children reached. A share cannot', 'p'),
        ('                               exceed 100%.', 'p'),
        ('  IMR, UFMR, ABR               capped at 1,000. These are rates per 1,000 births or', 'p'),
        ('                               women, so they may exceed 100, but not 1,000.', 'p'),
        ('  Any value below 0            set to 0.', 'p'),
        ('', None),
        ('No value is removed. All 5,987 rows are present with every indicator populated.', 'p'),
        ('', None),
        ('Also applied', 'h'),
        ('  - The 15 Pass/Fail and High/Low assessment columns are DROPPED. They were computed by', 'p'),
        ('    the source office against the uncapped values, and are not to be used.', 'p'),
        ('  - "#N/A" reference values are now blank.', 'p'),
        ('  - SORSOGON / PILAR / SAN ANTONIO resolved to PSGC 0506213048. All 5,987 barangays now', 'p'),
        ('    carry a PSGC code.', 'p'),
        ('  - Aligned to the final list: 5 Cavite barangays dropped, SUMISIP CENTRAL added from', 'p'),
        ('    the "2025 LIST" sheet (its ELCAC and Health Indicators columns are blank there).', 'p'),
        ('  - The FP CU column is absent — it was dropped from the reconciled file at source.', 'p'),
        ('    This dataset carries 12 indicators, not the original 13.', 'p'),
        ('', None),
        ('One column to decide on', 'h'),
        ('"Health Indicators" (0-7) is a count of how many health assessments passed, so it derives', 'p'),
        ('entirely from the dropped Pass/Fail columns and from uncapped values. It is retained here', 'p'),
        ('but is NOT derivable from the cleaned data. Drop it or recompute it before use.', 'p'),
        ('', None),
        ('Amber cells in "Cleaned data" are values changed by a rule.', 'p'),
        ('', None),
        ('Counts as produced', 'h'),
    ]
    r = 1
    for text, kind in lines:
        c = s.cell(row=r, column=1, value=text)
        c.font = TITLE if kind == 'title' else (BOLD if kind == 'h' else BODY)
        r += 1

    head(s, ['Indicator', 'Cap', 'Values capped', 'Max before', 'Max after'], row=r, height=30)
    r += 1
    src_max = {}
    for ci in INDICATORS:
        vals = [v for v in (rr[ci] for rr in allr[1:]) if isinstance(v, (int, float))]
        src_max[ci] = max(vals) if vals else 0
    rowsum = []
    for ci, label in list(CAP_100.items()) + list(CAP_1000.items()):
        cap = CAPS[ci]
        n = sum(1 for a in actions if a[1] == ci and a[4].startswith('Capped'))
        after = max((rr[ci] for rr in rows if isinstance(rr[ci], (int, float))), default=0)
        rowsum.append((label, cap, n, src_max[ci], after))
    for label, cap, n, before, after in sorted(rowsum, key=lambda x: -x[2]):
        for j, v in enumerate([label, cap, n, before, after], start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font, c.border = BODY, BOX
            if j > 1:
                c.number_format = '#,##0'
        r += 1
    for j, v in enumerate(['Total', '', n_cap100 + n_cap1000, '', ''], start=1):
        c = s.cell(row=r, column=j, value=v)
        c.font, c.border = BOLD, BOX
        if j == 3:
            c.number_format = '#,##0'
    r += 2
    for t in (f'{touched:,} of {len(rows):,} barangays had at least one value changed.',
              f'{n_na} "#N/A" cells blanked. {len(DROP_COLS)} Pass/Fail columns dropped.',
              f'PSGC attached for {len(rows) - psgc_missing:,} of {len(rows):,} barangays.',
              '0 values removed, 0 barangays dropped, 0 negatives found.'):
        s.cell(row=r, column=1, value=t).font = ITAL
        r += 1
    for col, w in zip('ABCDE', (30, 12, 16, 14, 14)):
        s.column_dimensions[col].width = w

    # ============================================================ Cleaned data
    s2 = out.create_sheet('Cleaned data')
    labels = ['PSGC'] + [str(HDR[i]) for i in keep_idx] + ['Values capped']
    head(s2, labels, height=42)
    act_by_cell = {(a[0], a[1]): a[4] for a in actions}
    for ri, r in enumerate(rows):
        xl = ri + 2
        c = s2.cell(row=xl, column=1, value=psgc_list[ri])
        c.font, c.border, c.number_format = BODY, BOX, '@'
        for j, ci in enumerate(keep_idx, start=2):
            c = s2.cell(row=xl, column=j, value=r[ci])
            c.font, c.border = BODY, BOX
            if isinstance(r[ci], (int, float)):
                c.number_format = '#,##0.##'
            if (ri, ci) in act_by_cell:
                c.fill = CAP_FILL
        n = sum(1 for ci in INDICATORS if (ri, ci) in act_by_cell)
        c = s2.cell(row=xl, column=len(keep_idx) + 2, value=n)
        c.font, c.border, c.number_format = BODY, BOX, '#,##0'
    s2.freeze_panes = 'F2'
    s2.auto_filter.ref = f'A1:{get_column_letter(len(labels))}{len(rows) + 1}'
    s2.column_dimensions['A'].width = 13
    for j, ci in enumerate(keep_idx, start=2):
        s2.column_dimensions[get_column_letter(j)].width = 20 if ci < 4 else 14
    s2.column_dimensions[get_column_letter(len(keep_idx) + 2)].width = 11

    # ============================================================ Actions log
    s3 = out.create_sheet('Actions log')
    head(s3, ['#', 'Region', 'Province', 'City / Municipality', 'Barangay', 'PSGC', 'Indicator',
              'Original value', 'Action taken', 'Value after', 'Source row'], height=34)
    for n, (ri, ci, label, orig, act, new) in enumerate(sorted(actions, key=lambda a: -a[3]), start=1):
        r = rows[ri]
        for j, v in enumerate([n, r[0], r[1], r[2], r[3], psgc_list[ri], label, orig, act, new,
                               ri + 2], start=1):
            c = s3.cell(row=n + 1, column=j, value=v)
            c.font, c.border = BODY, BOX
            if j == 6:
                c.number_format = '@'
            if j == 8:
                c.number_format, c.font, c.fill = '#,##0.##', BOLD, CAP_FILL
    s3.freeze_panes = 'A2'
    s3.auto_filter.ref = f'A1:K{len(actions) + 1}'
    for col, w in zip('ABCDEFGHIJK', (6, 40, 22, 26, 26, 13, 15, 14, 20, 12, 11)):
        s3.column_dimensions[col].width = w

    # ============================================================ Dropped columns
    s5 = out.create_sheet('Dropped columns')
    s5.sheet_view.showGridLines = False
    s5['A1'] = 'Columns removed from the cleaned dataset'
    s5['A1'].font = TITLE
    s5['A3'] = ('These are the source office\'s Pass/Fail and High/Low assessments. They were computed '
                'against the UNCAPPED values, so they no longer describe the data beside them.')
    s5['A3'].font = ITAL
    head(s5, ['Column in source', 'Kind'], row=5, height=24)
    for k, ci in enumerate(sorted(DROP_COLS)):
        for j, v in enumerate([str(HDR[ci]), 'Pass/Fail assessment'], start=1):
            c = s5.cell(row=6 + k, column=j, value=v)
            c.font, c.border = BODY, BOX
    s5.column_dimensions['A'].width = 28
    s5.column_dimensions['B'].width = 26

    # ============================================================ Provincial reference
    s6 = out.create_sheet('Provincial reference')
    labs = list(PROV_REF.values())
    head(s6, ['Region', 'Province', 'Barangays in list'] + labs + ['Adjusted to 100'], height=40)
    for n, rec in enumerate(prov_recs, start=2):
        base = [rec['region'], rec['province'] or '(HUC — no province)', rec['barangays']]
        for j, v in enumerate(base, start=1):
            c = s6.cell(row=n, column=j, value=v)
            c.font, c.border = BODY, BOX
            if j == 3:
                c.number_format = '#,##0'
        for k, lab in enumerate(labs):
            c = s6.cell(row=n, column=4 + k, value=rec['vals'].get(lab))
            c.font, c.border, c.number_format = BODY, BOX, '#,##0.00'
            if lab in rec['capped']:
                c.fill = CAP_FILL
        c = s6.cell(row=n, column=4 + len(labs),
                    value=', '.join(sorted(rec['capped'])) or '')
        c.font, c.border = BODY, BOX
    r0 = len(prov_recs) + 3
    for t, off in (('One row per province or HUC. These are the provincial comparators that '
                    'AO 2020-0023 §VI.A.2(d) tests each barangay against.', 0),
                   ('Amber = value adjusted down to 100. Pre-natal and Water are proportions, so a '
                    'provincial value above 100 is not possible.', 1),
                   ('IMR, UFMR and ABR are rates per 1,000 and are left as supplied — Samar\'s ABR '
                    'of 277 is legitimate.', 2),
                   ('FIC is NOT adjusted here: Ilocos Sur (102.15) and City of Butuan (101.00) '
                    'remain above 100. See the cleaning report.', 3),
                   ('Blank rows had no reference value in the source (#N/A).', 4)):
        s6.cell(row=r0 + off, column=1, value=t).font = ITAL
    s6.freeze_panes = 'D2'
    s6.auto_filter.ref = f'A1:{get_column_letter(4 + len(labs))}{len(prov_recs) + 1}'
    for col, w in zip(['A', 'B', 'C'], (40, 26, 12)):
        s6.column_dimensions[col].width = w
    for k in range(len(labs)):
        s6.column_dimensions[get_column_letter(4 + k)].width = 12
    s6.column_dimensions[get_column_letter(4 + len(labs))].width = 24

    # ============================================================ Summary
    s4 = out.create_sheet('Summary')
    s4.sheet_view.showGridLines = False
    s4['A1'] = 'Actions taken, by indicator'
    s4['A1'].font = TITLE
    s4['A3'] = (f'{len(rows):,} barangays retained  |  {n_cap100 + n_cap1000:,} values capped  |  '
                f'{n_neg:,} negatives  |  0 removed  |  {len(DROP_COLS)} columns dropped')
    s4['A3'].font = ITAL
    head(s4, ['Indicator', 'Cap', 'Values capped'], row=5, height=26)
    last = len(actions) + 1
    for k, (label, cap, n, _b, _a) in enumerate(sorted(rowsum, key=lambda x: -x[2])):
        row = 6 + k
        s4.cell(row=row, column=1, value=label).font = BODY
        s4.cell(row=row, column=2, value=cap).font = BODY
        s4.cell(row=row, column=3,
                value=f"=COUNTIFS('Actions log'!$G$2:$G${last},$A{row})")
        for j in range(1, 4):
            c = s4.cell(row=row, column=j)
            c.border = BOX
            if j > 1:
                c.number_format = '#,##0'
    tot = 6 + len(rowsum)
    s4.cell(row=tot, column=1, value='Total').font = BOLD
    s4.cell(row=tot, column=1).border = BOX
    s4.cell(row=tot, column=2).border = BOX
    c = s4.cell(row=tot, column=3, value=f'=SUM(C6:C{tot - 1})')
    c.font, c.number_format, c.border = BOLD, '#,##0', BOX
    for t, off in (('Water, Pre-natal, SBA and FIC are coverage percentages, capped at 100.', 2),
                   ('IMR, UFMR and ABR are rates per 1,000, capped at 1,000 — they may exceed 100.', 3),
                   ('No value was removed; all 5,987 barangays keep every indicator.', 4),
                   ('Pass/Fail columns dropped — see the "Dropped columns" sheet.', 5)):
        s4.cell(row=tot + off, column=1, value=t).font = ITAL
    for col, w in zip('ABC', (30, 12, 16)):
        s4.column_dimensions[col].width = w

    out.move_sheet('Summary', offset=-4)
    out.save(out_path)
    print('written:', out_path)

    if data_csv:
        # Same content as the "Cleaned data" sheet, in the form a loader reads:
        # ingestion/ingest_uuc_phc.py builds the seed migration from this file, so the
        # committed CSV — not the styled .xlsx — is what the database is reproducible from.
        act_cells = {(a[0], a[1]) for a in actions}
        # Which indicators were capped on each row, not just how many. A capped value is
        # indistinguishable from a genuine one once rendered -- 886 Water and 456 FIC values now
        # read as exactly 100% -- so the dashboard needs the flag per indicator to mark them
        # individually. `values_capped` (the count) stays for continuity with the workbook.
        capped_by_row = {}
        for ri, ci, label, _orig, action, _new in actions:
            if action.startswith('Capped'):
                capped_by_row.setdefault(ri, set()).add(csv_name(label))
        with open(data_csv, 'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(['psgc'] + [csv_name(HDR[i]) for i in keep_idx]
                       + ['values_capped', 'capped_indicators'])
            for ri, r in enumerate(rows):
                n = sum(1 for ci in INDICATORS if (ri, ci) in act_cells)
                w.writerow([psgc_list[ri]]
                           + ['' if r[ci] is None else r[ci] for ci in keep_idx]
                           + [n, '|'.join(sorted(capped_by_row.get(ri, ())))])
        print(f'written: {data_csv} ({len(rows):,} barangays, {len(keep_idx) + 2} columns)')

    if ref_csv:
        with open(ref_csv, 'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(['region', 'province', 'barangays_in_list']
                       + [f'ref_{l.lower().replace("-", "_")}' for l in labs]
                       + ['adjusted_to_100'])
            for rec in prov_recs:
                w.writerow([rec['region'], rec['province'], rec['barangays']]
                           + [rec['vals'].get(l, '') for l in labs]
                           + ['|'.join(sorted(rec['capped']))])
        print('written:', ref_csv, f'({len(prov_recs)} provinces, {n_prov_capped} values adjusted)')


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--src', default='ingestion/data/GIDA reconciled data.xlsx')
    ap.add_argument('--psgc-src', default='ingestion/data/Submissions_UUA_2025_filled_1.xlsx')
    ap.add_argument('--out', default='UUC_PHC_2025_cleaned.xlsx')
    ap.add_argument('--ref-csv', default='ingestion/data/uuc_phc_2025_provincial_reference.csv',
                    help='where to write the one-row-per-province reference table')
    ap.add_argument('--data-csv', default='ingestion/data/uuc_phc_2025_cleaned.csv',
                    help='where to write the machine-readable cleaned extract that '
                         'ingestion/ingest_uuc_phc.py loads')
    args = ap.parse_args()
    build(args.src, args.psgc_src, args.out, args.ref_csv, args.data_csv)
