#!/usr/bin/env python3
"""xlsx -> Postgres ingestion for PSA census population (E4.2).

Loads two PSA "Table B — Population ... by Province, City, and Municipality, By Region"
workbooks into agg_population (long format: one row per source × geo × census year):

  ingestion/data/psa_popcen_2024_tableB.xlsx  -> census_year 2024 (dataset psa-popcen-2024)
  ingestion/data/psa_cph_2020_tableB.xlsx     -> census_year 2020 (dataset psa-cph-2020)

Unlike ingest_stepzero.py, these PSA workbooks carry geography *names*, not PSGC codes,
so every row is name-matched to dim_geo (the join target of truth, post-NIR). Matching is
province-scoped: province names are globally unique, which disambiguates the ~200 duplicate
city/municipality names (e.g. "QUEZON" exists in four provinces). Region/province subtotal
rows are recognised structurally (their name matches a known region/province) and used only
as context, never as leaves — national/region/province figures are rolled up from the matched
city/municipality leaves via dim_geo's own post-NIR parentage. Rolling up from leaves (rather
than trusting each file's printed subtotals) is what makes the CPH 2020 numbers — published on
the pre-NIR vintage, with Negros under Regions VI/VII — reconcile onto dim_geo's post-NIR
Region XVIII automatically.

Known reconciliation residuals (documented, not silently dropped — the 1.6 discipline; see
docs/PSGC_CROSSWALK.md / docs/DECISIONS.md):
  * Manila is one census row but 16 district city/municipalities in dim_geo; its census total
    is stored at the Manila province node (geo_code 13806) and the districts are left with no
    census population (fall back to their parent, like agg_training's barangay gap).
  * The 2022 Maguindanao del Norte/del Sur split post-dates CPH 2020: its municipalities are
    matched within BARMM by name where unambiguous; the few genuinely ambiguous ones
    (e.g. PARANG, also in Sulu) are reported unmatched in the QA report.
  * A handful of newly-created / renamed LGUs absent from dim_geo are reported, not inserted
    with a dangling FK.

Grain: national -> region -> province -> citymun (these PSA releases stop at city/municipality;
there is no barangay population, so barangay-level per-capita falls back to citymun).

Run modes mirror ingest_stepzero.py:
  python ingestion/ingest_population.py --database-url "$DATABASE_URL"
  python ingestion/ingest_population.py --emit-sql-dir ingestion/_sql_batches_population
  python ingestion/ingest_population.py --verify        # offline: print reconciliation only
"""

import argparse
import json
import re
from pathlib import Path

import pandas as pd

from ingest import batched, insert_statement

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA = REPO_ROOT / "ingestion" / "data"
PARQUET_PATH = DATA / "dataset.parquet"
QA_REPORT_PATH = REPO_ROOT / "ingestion" / "_qa_report_population.json"

TABLE = "agg_population"
COLUMNS = ["dataset_id", "geo_code", "geo_level", "census_year", "population"]
BATCH_SIZE = 450

# One source workbook per census year. valcol = 0-indexed column holding that year's count
# (name in col 0; PGR columns follow the population columns). Verified against each file's
# header row: POPCEN cols = 2010/2015/2020/2024; CPH cols = 2000/2010/2015/2020.
SOURCES = [
    {"slug": "psa-popcen-2024", "file": "psa_popcen_2024_tableB.xlsx", "year": 2024, "valcol": 5,
     "xcheck_col": 4, "published_national": 112_729_484},
    {"slug": "psa-cph-2020", "file": "psa_cph_2020_tableB.xlsx", "year": 2020, "valcol": 5,
     "published_national": 109_035_343},
]

MANILA_NODE = "13806"           # dim_geo province-grain node "CITY OF MANILA (HUC)"
NIR_PROVINCES = {"NEGROS OCCIDENTAL", "NEGROS ORIENTAL", "SIQUIJOR"}
# Text rows (notes / sources / column headers) that must never be read as data leaves.
SKIP_SUBSTR = (
    "REGION PROVINCE AND CITY", "REGION PROVINCE AND HIGHLY", "POPULATION AND ANNUAL",
    "TOTAL POPULATION", "GROWTH RATE", "CENSUS OF POPULATION", "PHILIPPINE STATISTICS",
    "NOTE", "SOURCE", "EXCLUDES", "INCLUDES", "DATED", "ACCORDANCE",
    "PHILIPPINES", "NATIONAL CAPITAL REGION:", "BANGSAMORO AUTONOMOUS REGION",
)
ABBREV = {"GEN": "GENERAL", "PRES": "PRESIDENT", "STO": "SANTO", "STA": "SANTA"}
# Census spelling -> dim_geo spelling, for the handful that differ by one letter and would
# otherwise be lost. Documented reconciliations (docs/DECISIONS.md), not silent guesses.
SPELLING_FIXUPS = {
    "BALIUAG": "BALIWAG",              # dim_geo spells it BALIWAG (Bulacan)
    "PIO V CORPUS": "PIO V CORPUZ",    # dim_geo spells it CORPUZ (Masbate)
    "LEON T POSTIGO": "BACUNGAN",      # renamed municipality; dim_geo carries the old name (Zamb. del Norte)
    "DR JOSE P RIZAL": "RIZAL",        # dim_geo carries the short form (Palawan)
}


def norm(s):
    """Uppercase, drop footnote markers/asterisks/punctuation, expand PSA abbreviations."""
    if s is None:
        return ""
    s = str(s).strip().upper().replace("Ñ", "N")
    s = re.sub(r"\*+", " ", s)                 # footnote asterisks
    s = re.sub(r"\s+\d+\s*$", "", s)           # trailing footnote digit ("CITY OF BALIWAG 1")
    s = re.sub(r"[.,\-]", " ", s)              # punctuation + hyphen ("DATU SAUDI-AMPATUAN")
    s = re.sub(r"\s+", " ", s).strip()
    s = " ".join(ABBREV.get(w, w) for w in s.split())
    for a, b in SPELLING_FIXUPS.items():       # documented spelling reconciliations
        if s == a:
            return b
    return s.strip()


def prov_key(name):
    """Province-header match key: normalized, trailing parenthetical / 'excluding …'
    stripped ('SAMAR (WESTERN SAMAR)' -> 'SAMAR'), but the 'CITY OF …' prefix is kept
    so a provincial-capital city (e.g. 'CITY OF CEBU') is never mistaken for its province."""
    n = norm(name)
    return re.sub(r"\s*\(.+\)\s*$", "", n).strip()


def looks_like_city(name):
    n = norm(name)
    return n.startswith("CITY OF ") or n.endswith(" CITY")


def variants(name):
    """Candidate match keys in **priority order** (most specific first): full name, the
    'X CITY' -> 'CITY OF X' reorder, both halves of a 'NEWNAME (OLDNAME)' rename, then the
    class-stripped core. Ordered (not a set) so matching is deterministic — a set would let
    Python's randomised string hashing pick a different colliding code between runs."""
    n = norm(name)
    ordered = [n]
    m = re.match(r"^(.+?)\s+CITY$", n)         # "CALBAYOG CITY" -> "CITY OF CALBAYOG"
    if m:
        ordered.append(f"CITY OF {m.group(1)}")
    p = re.match(r"^(.*?)\s*\((.+)\)\s*$", n)   # "LICUAN-BAAY (LICUAN)"
    if p:
        for half in (p.group(1), p.group(2)):
            ordered.append(norm(half))
    core = re.sub(r"^(CITY OF|MUNICIPALITY OF)\s+", "", n)
    core = re.sub(r"\s*\((CAPITAL|CAPITAL CITY|HUC|ICC|CC)\)\s*$", "", core).strip()
    ordered.append(core)
    if m:
        ordered.append(m.group(1))
    if p:
        for half in (p.group(1), p.group(2)):
            ordered.append(re.sub(r"^(CITY OF|MUNICIPALITY OF)\s+", "", norm(half)))
    seen, res = set(), []
    for v in ordered:
        if v and v not in seen:
            seen.add(v)
            res.append(v)
    return res


def load_reference():
    pq = pd.read_parquet(PARQUET_PATH, columns=[
        "REGION CODE", "REGION NAME", "PROVINCE CODE", "PROVINCE NAME",
        "CITY/MUN CODE", "CITY/MUN NAME"])
    pq["REGION CODE"] = pq["REGION CODE"].astype(str).str.zfill(2)
    pq["PROVINCE CODE"] = pq["PROVINCE CODE"].astype(str).str.zfill(5)
    pq["CITY/MUN CODE"] = pq["CITY/MUN CODE"].astype(str).str.zfill(7)
    cm = pq.drop_duplicates("CITY/MUN CODE").copy()

    ref = {
        "cm": cm,
        # citymun_code -> (province_code, region_code), dim_geo's post-NIR parentage
        "parent": {r["CITY/MUN CODE"]: (r["PROVINCE CODE"], r["REGION CODE"]) for _, r in cm.iterrows()},
        "region_names": {norm(r): c for c, r in
                         pq.drop_duplicates("REGION CODE")[["REGION CODE", "REGION NAME"]].values},
        # province match-variant -> region_code (to region-scope province-header detection)
        "prov_region": {},
        "by_prov": {},   # (province_variant, citymun_variant) -> code
        "by_region": {}, # (region_code, citymun_variant) -> {codes}
        "all_codes": set(cm["CITY/MUN CODE"]),
    }
    for _, r in pq.drop_duplicates("PROVINCE CODE")[["PROVINCE NAME", "REGION CODE"]].iterrows():
        ref["prov_region"][prov_key(r["PROVINCE NAME"])] = r["REGION CODE"]
        ref["prov_region"][norm(r["PROVINCE NAME"])] = r["REGION CODE"]
    for _, r in cm.iterrows():
        rc = r["REGION CODE"]
        pkeys = {prov_key(r["PROVINCE NAME"]), norm(r["PROVINCE NAME"])}
        for pv in pkeys:
            for cv in variants(r["CITY/MUN NAME"]):
                ref["by_prov"].setdefault((pv, cv), r["CITY/MUN CODE"])
        for cv in variants(r["CITY/MUN NAME"]):
            ref["by_region"].setdefault((rc, cv), set()).add(r["CITY/MUN CODE"])
    return ref


def sheet_region_code(rows, ref):
    for row in rows[:14]:
        if row and row[0]:
            n = norm(row[0])
            for rn, rc in ref["region_names"].items():
                if rn and (rn == n or (len(rn) > 4 and rn in n)):
                    return rc
    return None


def is_region(n, ref):
    return any(rn == n or (len(rn) > 4 and rn in n and n in rn) for rn in ref["region_names"])


def next_leaf_name(rows, i, valcol, ref):
    """First data-leaf name after row i (skips blanks, notes, region rows)."""
    for j in range(i + 1, len(rows)):
        row = rows[j]
        raw = row[0] if row else None
        if not raw or not str(raw).strip():
            continue
        n = norm(raw)
        if any(s in n for s in SKIP_SUBSTR) or is_region(n, ref):
            continue
        val = row[valcol] if len(row) > valcol else None
        if isinstance(val, (int, float)):
            return raw
    return None


def is_province_header(raw, i, rows, valcol, allowed_prov, cur_prov, ref):
    """A province name is a header (not a same-named municipality) only when the next data
    leaf is one of *its own* municipalities — the reliable disambiguator for province=town
    name collisions (RIZAL-the-province vs RIZAL-the-Laguna-town)."""
    if looks_like_city(raw):
        return None
    pk = prov_key(raw)
    if pk not in allowed_prov and norm(raw) not in allowed_prov:
        return None
    if pk == cur_prov:
        return None   # already inside this province -> this is its eponymous town (BULACAN-in-Bulacan)
    nxt = next_leaf_name(rows, i, valcol, ref)
    if nxt is not None and not any(ref["by_prov"].get((pk, v)) for v in variants(nxt)):
        return None   # the following town isn't in this province -> this row is itself a town
    return pk


def parse_file(path, valcol, ref):
    """Return {citymun_code_or_manila_node: population} leaves + list of unmatched rows."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    leaves, unmatched = {}, []
    for sh in wb.sheetnames:
        rows = list(wb[sh].iter_rows(values_only=True))
        rc_sheet = sheet_region_code(rows, ref)
        allowed_prov = {pv for pv, r in ref["prov_region"].items() if r == rc_sheet} | NIR_PROVINCES
        cur_prov = None
        for i, row in enumerate(rows):
            raw = row[0] if row else None
            if not raw or not str(raw).strip():
                continue
            n = norm(raw)
            if any(s in n for s in SKIP_SUBSTR):
                continue
            if is_region(n, ref):
                cur_prov = None
                continue
            hdr = is_province_header(raw, i, rows, valcol, allowed_prov, cur_prov, ref)
            if hdr is not None:
                cur_prov = hdr
                continue
            val = row[valcol] if len(row) > valcol else None
            if not isinstance(val, (int, float)):
                continue
            code = None
            vs = variants(raw)
            if cur_prov:
                for v in vs:
                    code = ref["by_prov"].get((cur_prov, v))
                    if code:
                        break
            if not code and rc_sheet:
                for v in vs:
                    cand = ref["by_region"].get((rc_sheet, v))
                    if cand and len(cand) == 1:
                        code = next(iter(cand))
                        break
            if not code and "CITY OF MANILA" in n:
                code = MANILA_NODE
            if code:
                leaves[code] = int(val)
            else:
                unmatched.append({"sheet": sh, "province": cur_prov, "name": str(raw), "value": int(val)})
    return leaves, unmatched


def rollup_rows(leaves, ref, dataset_slug, year):
    """Emit agg_population rows at every dim_geo level, rolled up from matched leaves."""
    prov, reg = {}, {}
    national = 0
    rows = []
    for code, v in leaves.items():
        national += v
        if len(code) == 5:                       # Manila province node (province grain)
            prov[code] = prov.get(code, 0) + v
            reg[code[:2]] = reg.get(code[:2], 0) + v
            rows.append(_row(dataset_slug, code, "province", year, v))
        else:
            rows.append(_row(dataset_slug, code, "citymun", year, v))
            pc, rc = ref["parent"].get(code, (None, None))
            if pc:
                prov[pc] = prov.get(pc, 0) + v
            if rc:
                reg[rc] = reg.get(rc, 0) + v
    for pc, v in prov.items():
        if len(pc) == 5 and pc != MANILA_NODE:   # Manila node row already emitted above
            rows.append(_row(dataset_slug, pc, "province", year, v))
    for rc, v in reg.items():
        rows.append(_row(dataset_slug, rc, "region", year, v))
    rows.append(_row(dataset_slug, "PH", "national", year, national))
    return rows, national, reg


def _row(dataset_slug, geo_code, geo_level, year, population):
    return {
        "dataset_slug": dataset_slug,   # resolved to dataset_id at emit time
        "geo_code": geo_code,
        "geo_level": geo_level,
        "census_year": year,
        "population": population,
    }


def sql_for_rows(rows):
    """INSERT ... SELECT with dataset_id resolved via a slug subselect (self-contained and
    idempotent regardless of the generated dataset_id). All rows in a batch share one slug,
    so the slug is joined once rather than per row, keeping the statement compact."""
    slug = rows[0]["dataset_slug"]
    assert all(r["dataset_slug"] == slug for r in rows), "batch must be single-dataset"
    body = ",".join(
        f"('{r['geo_code']}','{r['geo_level']}',{r['census_year']},{r['population']})" for r in rows)
    return (
        f"INSERT INTO {TABLE} (dataset_id, geo_code, geo_level, census_year, population)\n"
        f"SELECT d.dataset_id, v.geo_code, v.geo_level, v.census_year, v.population\n"
        f"FROM (VALUES {body}) AS v(geo_code, geo_level, census_year, population)\n"
        f"CROSS JOIN (SELECT dataset_id FROM dim_dataset WHERE slug = '{slug}') d\n"
        f"ON CONFLICT (dataset_id, geo_code, geo_level, census_year) "
        f"DO UPDATE SET population = EXCLUDED.population;\n")


def selftest():
    """Guard the subtle name-normalisation and matching helpers (no DB / xlsx needed)."""
    assert norm("PEÑABLANCA") == "PENABLANCA"
    assert norm("CITY OF BALIWAG 1") == "CITY OF BALIWAG"          # trailing footnote digit
    assert norm("DATU SAUDI-AMPATUAN") == "DATU SAUDI AMPATUAN"    # hyphen -> space
    assert norm("GEN. MARIANO ALVAREZ") == "GENERAL MARIANO ALVAREZ"  # abbreviation
    assert norm("BALIUAG") == "BALIWAG" and norm("PIO V. CORPUS") == "PIO V CORPUZ"  # fixups
    assert prov_key("SAMAR (WESTERN SAMAR)") == "SAMAR"            # province paren strip
    assert prov_key("BASILAN (excluding CITY OF ISABELA)") == "BASILAN"
    assert looks_like_city("CITY OF CEBU") and looks_like_city("CALBAYOG CITY")
    assert not looks_like_city("CEBU")                            # the province, not the city
    assert "CITY OF CALBAYOG" in variants("CALBAYOG CITY")        # "X CITY" -> "CITY OF X"
    assert "LICUAN" in variants("LICUAN-BAAY (LICUAN)")           # rename paren half
    print("ingest_population selftest: OK")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--database-url")
    ap.add_argument("--emit-sql-dir")
    ap.add_argument("--verify", action="store_true", help="offline: print reconciliation only")
    ap.add_argument("--selftest", action="store_true", help="run helper assertions and exit")
    args = ap.parse_args()

    if args.selftest:
        selftest()
        return

    ref = load_reference()
    qa = {"grain": "national->region->province->citymun (no barangay in source)", "sources": []}
    all_rows = []

    for src in SOURCES:
        leaves, unmatched = parse_file(DATA / src["file"], src["valcol"], ref)
        rows, national, reg = rollup_rows(leaves, ref, src["slug"], src["year"])
        all_rows += rows
        delta = national - src["published_national"]
        entry = {
            "slug": src["slug"], "year": src["year"], "file": src["file"],
            "leaves_matched": len(leaves),
            "unmatched_count": len(unmatched),
            "national_rollup": national,
            "national_published": src["published_national"],
            "national_delta": delta,
            "national_delta_pct": round(100.0 * delta / src["published_national"], 3),
            "regions_covered": len(reg),
            "citymun_uncovered": len(ref["all_codes"] - {c for c in leaves if len(c) == 7}),
            "unmatched_sample": unmatched,
        }
        if "xcheck_col" in src:
            xleaves, _ = parse_file(DATA / src["file"], src["xcheck_col"], ref)
            entry["xcheck_2020_from_this_file"] = sum(xleaves.values())
        qa["sources"].append(entry)
        print(f"[{src['slug']}] leaves={len(leaves)} national={national:,} "
              f"published={src['published_national']:,} delta={delta:+,} "
              f"({entry['national_delta_pct']:+}%) regions={len(reg)} unmatched={len(unmatched)}")
        for u in unmatched:
            print("   UNMATCHED:", u["sheet"], u["province"], repr(u["name"]), u["value"])

    qa["total_rows"] = len(all_rows)
    QA_REPORT_PATH.write_text(json.dumps(qa, indent=2, default=str))
    print(f"\nTotal agg_population rows: {len(all_rows)}  (QA -> {QA_REPORT_PATH.name})")

    if args.verify:
        return
    # Batch per dataset — sql_for_rows() requires a single-dataset batch.
    def per_dataset_batches():
        for src in SOURCES:
            rows = [r for r in all_rows if r["dataset_slug"] == src["slug"]]
            for chunk in batched(rows, BATCH_SIZE):
                yield src["slug"], chunk

    if args.emit_sql_dir:
        out = Path(args.emit_sql_dir)
        out.mkdir(parents=True, exist_ok=True)
        for i, (slug, chunk) in enumerate(per_dataset_batches(), 1):
            (out / f"{i:04d}_{slug}_{TABLE}.sql").write_text(sql_for_rows(chunk))
        print(f"Wrote SQL batches to {out}")
    elif args.database_url:
        import psycopg2
        conn = psycopg2.connect(args.database_url)
        try:
            with conn, conn.cursor() as cur:
                for _, chunk in per_dataset_batches():
                    cur.execute(sql_for_rows(chunk))
        finally:
            conn.close()
        print("Loaded via psycopg2.")
    else:
        ap.error("pass --verify, --emit-sql-dir, or --database-url")


if __name__ == "__main__":
    main()
