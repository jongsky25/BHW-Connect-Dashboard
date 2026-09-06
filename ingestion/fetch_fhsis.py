"""Pull the 2025 FHSIS annual workbooks this repo loads out of DOH's public Drive archive.

Source: `https://bit.ly/FHSISPHSannualreports` → Drive folder `16z6srVbGODqmgGHU4_Qg1oDBOglqp_XG`,
owned by `fhsisreports@doh.gov.ph`, readable with no login. Six subfolders (Annual / Quarterly /
Monthly, each in Excel and PDF); this script reads the **Annual Excel 2025** folder
(`1x_aBwfWDGyylC75RA1CTQM6JGwiwLINQ`) and nothing else, because plan F1 loads Annual only.

**Why this script is committed at all.** The archive is mutable — DOH keeps adding to and
re-touching it, and `Demographic_2025_EB_Final.xlsx` carried a Drive modified date of
2026-08-24 when this was written, i.e. eight months after the year it reports on. What the loader
publishes is therefore a *snapshot*, and a snapshot is only reproducible if the exact file ids and
their modified dates are recorded. That is the same reason `ingestion/redact_nhfr_source.py` is
committed: the retrieval, not just the result, is part of the provenance. The manifest this writes
is the diff target for a future pull (plan Decision 9).

**The raw workbooks are not committed; the manifest is.** Decision 8 names the long-format cleaned
CSVs plus `docs/FHSIS_2025_CLEANING_REPORT.md` as this dataset's reproducible committed artefacts,
so `ingestion/data/fhsis_2025/` is gitignored and `_manifest.json` inside it is the one file that
is checked in — file id, Drive modified date, byte size and both digests per workbook. An operator
re-running this script gets either the identical manifest (the archive has not moved) or a visible
diff (it has).

Two mechanics the archive forces, both found by trying it rather than from documentation:

  - **A browser User-Agent is required.** Bare curl/urllib gets 403 from `doh.gov.ph`; Drive
    itself is content with any UA, but the ten program-area entries in the 2025 folder are
    *shortcuts* to separately-shared folders, so a fetch has to follow redirects too. Both are
    set below rather than left to the caller.
  - **Two download endpoints, because the archive mixes two file kinds.** Real `.xlsx` uploads come
    from `uc?export=download`; three of the Envi files are native Google Sheets, which have no
    stored bytes and must be rendered through `/export?format=xlsx`. Which is which is a property
    of the file, recorded per entry below.

The Drive modified date comes from the download response's `Last-Modified` header, not from the
folder listing. The listing renders it as "Aug 24" — no year, no time — while the header gives
`Mon, 24 Aug 2026 07:11:43 GMT`, which is what a provenance record needs.

**The three native Sheets are the exception, twice over, and the manifest says so rather than
hiding it.** A Google Sheet has no stored bytes: `/export?format=xlsx` renders one on request.

  1. *No modified date.* The export sends no `Last-Modified` at all, so the only date Drive
     exposes for them without authenticating is the folder listing's day-precision cell
     ("Jul 30"). Those three carry `drive_modified_precision: "day"` and a date read from the
     listing; the six real uploads carry `"second"` and the header's timestamp.
  2. *No stable bytes.* Two exports of an **unchanged** sheet differ — verified by pulling
     `zod_nofml` twice in a row and getting two SHA-256s, with `sharedStrings.xml`, `styles.xml`
     and every `sheet*.xml` differing. The renderer is not deterministic, so a byte hash of a
     Sheet answers "was this rendered twice", not "did the data change".

So the manifest records **two** digests and the diff uses the right one. `sha256` is the raw
bytes, meaningful only for `kind == "xlsx"`. `content_digest` is a SHA-256 over every sheet's
*cell values* — stable across re-exports of an unchanged Sheet, and the field `--check` actually
compares. Hashing what the loader reads, rather than what the transport happened to produce, is
the only version signal that means the same thing for both file kinds.

Usage:
    python ingestion/fetch_fhsis.py                 # fetch all nine into ingestion/data/fhsis_2025/
    python ingestion/fetch_fhsis.py --area envi     # one program area
    python ingestion/fetch_fhsis.py --check         # re-read the archive, diff against the manifest
"""

import argparse
import datetime as dt
import hashlib
import json
import urllib.error
import urllib.request
from pathlib import Path

OUT_DIR_DEFAULT = "ingestion/data/fhsis_2025"
MANIFEST_NAME = "_manifest.json"

# The public archive, for the record and for the cleaning report's citation. Not fetched here —
# the file ids below were read out of these folders and are pinned so a run is reproducible.
ARCHIVE_FOLDER = "16z6srVbGODqmgGHU4_Qg1oDBOglqp_XG"
ANNUAL_EXCEL_2025_FOLDER = "1x_aBwfWDGyylC75RA1CTQM6JGwiwLINQ"
ARCHIVE_SHORTLINK = "https://bit.ly/FHSISPHSannualreports"

# The Envi program-area folder, resolved from its shortcut in the 2025 listing. Named because the
# three Sheets inside it need their modified date read from this listing — see the docstring.
ENVI_FOLDER = "1oyu_5XhX7HyuQsXqCQrVPXpuB7akPrQN"

# Chrome's UA. See the module docstring: the archive's own host 403s a bare client, and a fetch
# that works today from a browser and not from this script is a fetch that will be debugged twice.
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)

# The F1 slice: five program areas, Tier 1 (city/municipality, PSGC-keyed) only.
#
# Every id below was enumerated from the 2025 folder rather than guessed, and the enumeration is
# re-runnable — but the ids are pinned rather than re-discovered on each run, because a fetch that
# silently picks up a *different* file when DOH renames one is exactly the failure this dataset's
# mutability makes likely. `--check` is the deliberate re-enumeration path.
#
# `kind` is 'xlsx' for a real uploaded workbook and 'sheet' for a native Google Sheet, which has no
# stored bytes and is rendered on demand — the three Envi files are the latter.
WORKBOOKS = [
    {
        "key": "demographics",
        "area": "demographics",
        "file_id": "1Ef3-GV3M7YWPhOJLZWY_qukLxROLL6JM",
        "kind": "xlsx",
        "filename": "Demographic_2025_EB_Final.xlsx",
        "drive_folder": "Demographics",
        "what": "Population/household projections, barangay/RHU/BHS counts, and the health "
                "workforce split LGU-hired / DOH-hired. The BHW columns in this workbook are "
                "dropped by clean_fhsis.py and reach no table — plan Decision 2.",
    },
    {
        "key": "immunization_fic_cic",
        "area": "immunization",
        "file_id": "1HX57uAX9aktbEmgbt81RhrqsdfJxqV5T",
        "kind": "xlsx",
        "filename": "9 10 FIC _ CIC_MunCity EB 2025_FInal Annual 2025_NF.xlsx",
        "drive_folder": "Childcare / Immunization - Final 2025",
        "what": "Fully and completely immunised children. The antigen-by-antigen files in the "
                "same folder (BCG, DPT, OPV, IPV, PCV, MMR) are deferred, per plan Decision 3.",
    },
    {
        "key": "maternal_anc8",
        "area": "maternal",
        "file_id": "1M0rggBNSyW6eNikPvwvu4CrABIFmTtIY",
        "kind": "xlsx",
        "filename": "1. 8ANC_2025_PH_nofml.xlsx",
        "drive_folder": "Maternal Care / Prenatal care",
        "what": "Pregnant women with at least 8 antenatal care visits.",
    },
    {
        "key": "maternal_anc4",
        "area": "maternal",
        "file_id": "1BUu8lt-AsQUDMJROjaxCLgYL6onAhq4A",
        "kind": "xlsx",
        "filename": "1. 4ANC_2025_EB_1_nofml.xlsx",
        "drive_folder": "Maternal Care / 4anc and 2pnc",
        "what": "Pregnant women with at least 4 antenatal care visits — the criterion-(d) "
                "indicator /uuc-phc already publishes for its 5,987 listed barangays.",
    },
    {
        "key": "maternal_pnc2",
        "area": "maternal",
        "file_id": "11tdkuWcifmgmP1SEuH-3LwHacdHQaEq0",
        "kind": "xlsx",
        "filename": "2. 2PNC_2025_EB_nofml.xlsx",
        "drive_folder": "Maternal Care / 4anc and 2pnc",
        "what": "Postpartum women with at least 2 postnatal care visits.",
    },
    {
        "key": "envi_water",
        "area": "envi",
        "file_id": "1T0ToCiqUUucP8bZwvOH3Lqpf_nLGR9sNP7Yucw_X840",
        "kind": "sheet",
        "filename": "water_nofml.xlsx",
        "drive_folder": "Envi",
        "parent_folder_id": ENVI_FOLDER,
        "what": "Households with basic and with safely-managed drinking water.",
    },
    {
        "key": "envi_sanitation",
        "area": "envi",
        "file_id": "10iHGNT3D6D8YPr-AsbMp3dkvJHzcrm3CNyJnx8mz-uo",
        "kind": "sheet",
        "filename": "sanitation_nofml.xlsx",
        "drive_folder": "Envi",
        "parent_folder_id": ENVI_FOLDER,
        "what": "Households with basic sanitation facilities.",
    },
    {
        "key": "envi_zod",
        "area": "envi",
        "file_id": "14exmLYtq0z0QNdtVRiT7lCTwG_pm1AiNRvl8IWDF6l8",
        "kind": "sheet",
        "filename": "zod_nofml.xlsx",
        "drive_folder": "Envi",
        "parent_folder_id": ENVI_FOLDER,
        "what": "Zero open defecation — barangays and households declared ZOD.",
    },
    {
        "key": "tb",
        "area": "tb",
        "file_id": "1T0b9zhP7pj8qOZ15U6-aYiVXK2Nsfdqc",
        "kind": "xlsx",
        "filename": "FINAL Tuberculosis_EB_2025.xlsx",
        "drive_folder": "Infectious / TB",
        "what": "The TB cascade: notified, bacteriologically confirmed, TPT, treatment success.",
    },
]

AREAS = sorted({w["area"] for w in WORKBOOKS})


def download_url(entry: dict) -> str:
    """The endpoint for one entry. See the module docstring on the two file kinds."""
    if entry["kind"] == "sheet":
        return (
            f"https://docs.google.com/spreadsheets/d/{entry['file_id']}/export?format=xlsx"
        )
    return f"https://drive.google.com/uc?export=download&id={entry['file_id']}"


def fetch(url: str) -> tuple[bytes, str | None]:
    """GET with a browser UA, following redirects. Returns (body, Last-Modified header).

    urllib follows redirects by default and carries the headers with it, which is what the
    shortcut→shared-folder hop and Drive's own 303 to `drive.usercontent.google.com` both need.
    """
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=300) as resp:
        return resp.read(), resp.headers.get("Last-Modified")


def http_date_to_iso(value: str | None) -> str | None:
    """`Mon, 24 Aug 2026 07:11:43 GMT` -> `2026-08-24T07:11:43+00:00`.

    Returned as None rather than guessed if Drive stops sending the header: a provenance record
    with an invented date is worse than one that says the date is unknown.
    """
    if not value:
        return None
    try:
        return (
            dt.datetime.strptime(value, "%a, %d %b %Y %H:%M:%S %Z")
            .replace(tzinfo=dt.timezone.utc)
            .isoformat()
        )
    except ValueError:
        return None


def listing_modified_date(folder_id: str, file_id: str) -> str | None:
    """Day-precision modified date for one file, read from its Drive folder listing.

    The fallback for native Google Sheets, which send no `Last-Modified` (see the docstring).
    Drive's listing renders the date as "Jul 30" inside the last twelve months and "Jul 30, 2025"
    beyond them; both forms are parsed, and the bare form resolves to the most recent occurrence
    that is not in the future — which is what Drive itself means by it.

    Returns None on any failure. This is a best-effort enrichment of a provenance record, not a
    load-blocking step: the SHA-256 is what a future pull actually compares.
    """
    import html as html_mod
    import re

    url = f"https://drive.google.com/drive/folders/{folder_id}"
    try:
        body, _ = fetch(url)
    except (urllib.error.URLError, TimeoutError):
        return None
    page = body.decode("utf-8", errors="replace")

    for row in re.split(r'<tr data-selectable data-id="', page)[1:]:
        if not row.startswith(file_id):
            continue
        cells = [
            html_mod.unescape(c).strip()
            for c in re.sub(r"<[^>]+>", "\x01", row[: row.find("</tr>")]).split("\x01")
            if c.strip()
        ]
        for cell in cells:
            for fmt, has_year in (("%b %d, %Y", True), ("%b %d", False)):
                try:
                    parsed = dt.datetime.strptime(cell, fmt).date()
                except ValueError:
                    continue
                if has_year:
                    return parsed.isoformat()
                today = dt.date.today()
                year = today.year if parsed.replace(year=today.year) <= today else today.year - 1
                return parsed.replace(year=year).isoformat()
    return None


def content_digest(path: Path) -> str:
    """SHA-256 over every sheet's cell values, in sheet then row then column order.

    The version signal for this archive, for the reason the module docstring gives: a native
    Google Sheet re-renders to different bytes every time it is exported, so a raw byte hash
    reports a change on every pull. This hashes what the cleaner will actually read, so it moves
    when — and only when — a figure moves.

    `data_only=True` reads cached formula results rather than formulas, which is also what
    clean_fhsis.py reads, so the digest covers the same values the load does.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        digest = hashlib.sha256()
        for name in workbook.sheetnames:
            digest.update(name.encode("utf-8") + b"\x00")
            for row in workbook[name].iter_rows(values_only=True):
                cells = "\x1f".join("" if c is None else str(c) for c in row)
                digest.update(cells.encode("utf-8") + b"\x1e")
        return digest.hexdigest()
    finally:
        workbook.close()


def looks_like_xlsx(body: bytes) -> bool:
    """An .xlsx is a zip. Drive answers an unauthorised or interstitial request with HTML and a
    200, so the content type has to be checked from the bytes rather than trusted."""
    return body[:2] == b"PK"


def fetch_one(entry: dict, out_dir: Path) -> dict:
    url = download_url(entry)
    body, last_modified = fetch(url)
    if not looks_like_xlsx(body):
        raise SystemExit(
            f"{entry['key']}: {url} returned {len(body):,} bytes that are not a workbook "
            f"(starts {body[:40]!r}).\nThe file may have been unshared, or Drive served an "
            "interstitial — open the URL in a browser to see which."
        )

    path = out_dir / entry["filename"]
    path.write_bytes(body)

    modified = http_date_to_iso(last_modified)
    precision = "second" if modified else None
    if not modified and entry.get("parent_folder_id"):
        modified = listing_modified_date(entry["parent_folder_id"], entry["file_id"])
        precision = "day" if modified else None

    return {
        "key": entry["key"],
        "area": entry["area"],
        "file_id": entry["file_id"],
        "kind": entry["kind"],
        "filename": entry["filename"],
        "drive_folder": entry["drive_folder"],
        "what": entry["what"],
        "download_url": url,
        # The provenance pair plan Decision 9 requires. drive_modified_at is the archive's own
        # timestamp for the file; retrieved_at (manifest-level) is when this snapshot was taken.
        # The precision is recorded beside it because it is not the same for both file kinds.
        "drive_modified_at": modified,
        "drive_modified_precision": precision,
        "bytes": len(body),
        # Raw transport bytes. Meaningful for an uploaded .xlsx; for a Sheet it changes on every
        # export whether or not the data did — see the module docstring. Recorded anyway, because
        # for the six uploads it is the cheapest possible integrity check.
        "sha256": hashlib.sha256(body).hexdigest(),
        # The value-level digest --check compares. Stable for both file kinds.
        "content_digest": content_digest(path),
    }


def load_manifest(out_dir: Path) -> dict | None:
    path = out_dir / MANIFEST_NAME
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def diff_manifest(old: dict | None, new: dict) -> list[str]:
    """What moved in the archive since the committed manifest. Reported, never auto-accepted:
    a changed workbook means a re-clean, a re-load and a `bumpDatasetVersion` (Decision 9)."""
    if not old:
        return ["no previous manifest to compare against — this is the first pull"]
    before = {e["key"]: e for e in old["workbooks"]}
    after = {e["key"]: e for e in new["workbooks"]}
    changes = []
    for key in sorted(set(before) | set(after)):
        if key not in before:
            changes.append(f"{key}: new in this pull")
        elif key not in after:
            changes.append(f"{key}: no longer fetched")
        elif before[key].get("content_digest") != after[key].get("content_digest"):
            # content_digest, never sha256: the three Sheets re-render to new bytes on every pull,
            # so a byte comparison would cry wolf three times a run and train the operator to
            # ignore it.
            changes.append(
                f"{key}: CONTENT CHANGED — was {before[key].get('content_digest', '?')[:12]} "
                f"(modified {before[key]['drive_modified_at']}), now "
                f"{after[key]['content_digest'][:12]} (modified {after[key]['drive_modified_at']})"
            )
    return changes


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out-dir", default=OUT_DIR_DEFAULT, help="Where to write the workbooks")
    parser.add_argument(
        "--area", choices=AREAS, help="Fetch one program area only (default: all five)"
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Fetch and diff against the committed manifest without overwriting it",
    )
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wanted = [w for w in WORKBOOKS if not args.area or w["area"] == args.area]
    print(f"Fetching {len(wanted)} workbook(s) from the 2025 Annual Excel folder")
    print(f"  archive: {ARCHIVE_SHORTLINK} -> {ARCHIVE_FOLDER}")
    print(f"  folder:  {ANNUAL_EXCEL_2025_FOLDER}")
    print()

    fetched = []
    for entry in wanted:
        record = fetch_one(entry, out_dir)
        fetched.append(record)
        print(
            f"  {record['key']:<22} {record['bytes']:>9,} B  "
            f"modified {record['drive_modified_at'] or '(not stated)'}  {record['filename']}"
        )

    manifest = {
        "archive_folder": ARCHIVE_FOLDER,
        "annual_excel_2025_folder": ANNUAL_EXCEL_2025_FOLDER,
        "archive_shortlink": ARCHIVE_SHORTLINK,
        "retrieved_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat(),
        "workbooks": fetched,
    }

    previous = load_manifest(out_dir)
    if args.area or args.check:
        # A partial or check run must not rewrite a manifest that describes all nine.
        print("\nArchive diff:")
        for line in diff_manifest(previous, manifest):
            print(f"  - {line}")
        if args.check:
            return

    if not args.area:
        (out_dir / MANIFEST_NAME).write_text(
            json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
        )
        print(f"\nManifest written to {out_dir / MANIFEST_NAME}")


if __name__ == "__main__":
    main()
