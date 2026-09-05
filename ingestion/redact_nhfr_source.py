"""Produce the redacted copy of the NHFR export that this repository commits.

The public NHFR export carries personal contact details: of its 20,194 email addresses, 18,413
are free webmail accounts belonging to individual midwives, proprietors and staff rather than to
institutions. This repository is public, so the source workbook is committed **with those columns
removed** — the pipeline never reads them (see `clean_nhfr.py`), and there is no reason for them
to sit in a public git history.

What is removed, and nothing else:

    Landline Number, Landline Number 2, Fax Number,
    Email Address, Alternate Email Address, Official Website

Everything the published figures derive from — codes, names, types, ownership, bed capacity,
licensing, and the street/building address fields — is left exactly as the source printed it, so
`clean_nhfr.py` can still be re-run against the committed file and reproduce the extract.

Usage (the operator supplies the unredacted export they downloaded):

    python ingestion/redact_nhfr_source.py \
        --src ~/Downloads/nhfr_universe_as_of_September_2026.xlsx \
        --out ingestion/data/nhfr_universe_as_of_September_2026.xlsx
"""

import argparse

import openpyxl

REDACTED_COLUMNS = [
    "Landline Number",
    "Landline Number 2",
    "Fax Number",
    "Email Address",
    "Alternate Email Address",
    "Official Website",
]


def redact(src: str, out: str) -> None:
    wb = openpyxl.load_workbook(src)
    total_cleared = 0

    for ws in wb.worksheets:
        header = [cell.value for cell in ws[1]]
        targets = [i + 1 for i, name in enumerate(header) if name in REDACTED_COLUMNS]
        if not targets:
            continue
        for col in targets:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    cell.value = None
                    total_cleared += 1

    wb.save(out)
    print(f"Wrote {out}")
    print(f"  cleared {total_cleared:,} values across {len(REDACTED_COLUMNS)} columns:")
    for name in REDACTED_COLUMNS:
        print(f"    - {name}")
    print("  headers are kept, so the file's shape still matches what clean_nhfr.py expects.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--src", required=True, help="The unredacted export as downloaded")
    parser.add_argument("--out", required=True, help="Where to write the redacted copy")
    args = parser.parse_args()
    redact(args.src, args.out)


if __name__ == "__main__":
    main()
