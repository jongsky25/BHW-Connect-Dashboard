"""Turn the public NHFR export into the committed machine-readable extract the loader reads.

Source: `ingestion/data/nhfr_universe_as_of_September_2026.xlsx` — the DOH National Health
Facility Registry "universe" list, exported from the public site `nhfr.doh.gov.ph` and retrieved
2026-09-05. NHFR is a live registry rather than a periodic publication, so this file is a
*snapshot*: the dataset slug carries its month (`nhfr-2026-09`) and a later export is a new
version, not a correction of this one.

This mirrors `ingestion/clean_uuc_phc_indicators.py`: the cleaning step is separate from the load,
the loader never opens the workbook, and every action taken here is reported rather than applied
silently.

What this does, and why each one:

  - **Drops every contact and street-address column.** Of the 20,194 email addresses in the
    export, 18,413 (91%) are free webmail — gmail/yahoo/hotmail/outlook — i.e. the personal
    addresses of individual midwives and proprietors, not institutional contacts. The published
    tables are anon-readable over PostgREST under CC BY 4.0, so loading them would republish
    ~18,000 personal email addresses. `docs/BUILD_PLAN.md` pitfall P16 sets the precedent (the
    free-text training column "never leaves raw tables"), and no planned figure needs them. The
    barangay is the location granularity the dashboard actually renders.

  - **Strips the `"Licensing Status:"` prefix.** The source stores the field name inside the
    value ("Licensing Status:With License"), which is a source artefact, not a value.

  - **Derives `citymun_code`** by truncating the source's 10-digit city/municipality PSGC to
    `dim_geo`'s 7-digit width. Verified lossless: every city/municipality code in the export ends
    in three zeros, every province code in five, every region code in eight.

  - **Folds the two ownership sub-classification columns into one.** They are mutually exclusive
    by construction — a facility is Government or Private, never both — so two columns that are
    each mostly blank become one that is populated when the source knows the answer.

  - **Normalises `License Validity Date`** to ISO. The source writes M/D/YYYY or a real date cell.

Nothing is inferred, imputed or dropped row-wise: all 44,799 facilities reach the extract,
including the 108 with no barangay code (they keep a blank `barangay_code` and are joined at
city/municipality level instead — see `docs/DATASET_SCOPING.md` and the N1 plan).

Usage:
    python ingestion/clean_nhfr.py
"""

import argparse
import csv
import datetime as dt
from collections import Counter

import openpyxl

SRC_DEFAULT = "ingestion/data/nhfr_universe_as_of_September_2026.xlsx"
OUT_DEFAULT = "ingestion/data/nhfr_2026_09_cleaned.csv"

SHEET = "Worksheet"

# Source column index (0-based) -> what it is. The export's own header order, verified against the
# file rather than assumed; the loader re-checks the header before it trusts a single row.
SRC = {
    "facility_code": 0,
    "facility_code_short": 1,
    "facility_name": 2,
    "facility_major_type": 6,
    "facility_type": 7,
    "ownership_major": 8,
    "ownership_sub_gov": 9,
    "ownership_sub_private": 10,
    "region_name": 13,
    "region_psgc": 14,
    "province_name": 15,
    "province_psgc": 16,
    "citymun_name": 17,
    "citymun_psgc": 18,
    "barangay_name": 19,
    "barangay_psgc": 20,
    "service_capability": 28,
    "bed_capacity": 29,
    "licensing_status": 30,
    "license_validity_date": 31,
}

# The header the export must have, in order, for the indices above to mean what they say.
EXPECTED_HEADER = [
    "Health Facility Code",
    "Health Facility Code Short",
    "Facility Name",
    "Old Health Facility Name 1",
    "Old Health Facility Name 2",
    "Old Health Facility Name 3",
    "Facility Major Type",
    "Health Facility Type",
    "Ownership Major Classification",
    "Ownership Sub-Classification for Government facilities",
    "Ownership Sub-Classification for private facilities",
    "Street Name and #",
    "Building name and #",
    "Region Name",
    "Region PSGC",
    "Province Name",
    "Province PSGC",
    "City/Municipality Name",
    "City/Municipality PSGC",
    "Barangay Name",
    "Barangay PSGC",
    "Zip Code",
    "Landline Number",
    "Landline Number 2",
    "Fax Number",
    "Email Address",
    "Alternate Email Address",
    "Official Website",
    "Service Capability",
    "Bed Capacity",
    "Licensing Status",
    "License Validity Date",
]

# Columns deliberately not carried into the extract. Named rather than merely omitted so the
# exclusion is a stated decision a reader can audit, not an oversight.
EXCLUDED = {
    "Old Health Facility Name 1": "6.75% populated; superseded names are not rendered anywhere",
    "Old Health Facility Name 2": "0.73% populated",
    "Old Health Facility Name 3": "0.24% populated",
    "Street Name and #": "street address not needed at barangay granularity",
    "Building name and #": "street address not needed at barangay granularity",
    "Zip Code": "redundant with PSGC",
    "Landline Number": "personal/contact detail — see module docstring",
    "Landline Number 2": "personal/contact detail — see module docstring",
    "Fax Number": "personal/contact detail — see module docstring",
    "Email Address": "91% free webmail, i.e. personal addresses — see module docstring",
    "Alternate Email Address": "personal/contact detail — see module docstring",
    "Official Website": "0.60% populated; contact detail",
}

LICENSING_PREFIX = "Licensing Status:"

OUT_COLUMNS = [
    "facility_code",
    "facility_code_short",
    "facility_name",
    "facility_major_type",
    "facility_type",
    "ownership_major",
    "ownership_sub",
    "source_region_psgc",
    "source_region_name",
    "source_province_psgc",
    "source_province_name",
    "source_citymun_psgc",
    "source_citymun_name",
    "source_barangay_psgc",
    "source_barangay_name",
    "citymun_code",
    "barangay_code",
    "service_capability",
    "bed_capacity",
    "licensing_status",
    "license_validity_date",
]


def text(value) -> str:
    """Source cell -> trimmed string. None and whitespace-only both become ''."""
    if value is None:
        return ""
    return str(value).strip()


def psgc(value) -> str:
    """Source PSGC cell -> the 10-digit string the export stores.

    openpyxl hands back an int when a code has no leading zero (Regions 10-19), so the width is
    restored here rather than trusted. This is BUILD_PLAN pitfall P1 in its original form.
    """
    raw = text(value)
    if not raw:
        return ""
    if raw.endswith(".0"):  # a numeric cell read back as a float
        raw = raw[:-2]
    return raw.zfill(10)


def iso_date(value, problems: list, facility_code: str) -> str:
    """License validity date -> ISO. Unparseable values are reported, never guessed."""
    if value is None or text(value) == "":
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    raw = text(value)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return dt.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    problems.append(f"{facility_code}: unparseable License Validity Date {raw!r}")
    return ""


def bed_capacity(value, problems: list, facility_code: str) -> str:
    """Bed capacity -> integer string. Non-numeric is reported, not coerced to 0.

    Six of the largest hospitals write the figure with thousands separators ("1,200"), which is
    formatting rather than a different value, so the separator is removed before parsing. Anything
    still unparseable after that is a real disagreement and is reported.
    """
    raw = text(value).replace(",", "")
    if raw == "":
        return "0"
    try:
        return str(int(float(raw)))
    except ValueError:
        problems.append(f"{facility_code}: non-numeric Bed Capacity {text(value)!r}")
        return "0"


def ownership_sub(gov_sub: str, private_sub: str, ownership_major: str) -> str:
    """The two sub-classification columns folded into one, resolved by the major classification.

    They are *meant* to be mutually exclusive, and for 44,784 of the 44,799 facilities they are.
    Fifteen carry both — a Government/LGU rural health unit also tagged "Single Proprietorship",
    a Private corporate hospital also tagged "Local Government Unit". In every one of them the
    sub-classification that contradicts `Ownership Major Classification` is the stray value, so
    the rule is to honour the major classification and take the sub-classification belonging to
    it. The exceptions are counted and named in the cleaning report rather than silently resolved,
    on the precedent `docs/BUILD_PLAN.md` pitfall P15 sets for the honorarium flag/amount
    inconsistencies: reconcile by a stated rule, log the exceptions.
    """
    if ownership_major == "Government":
        return gov_sub
    if ownership_major == "Private":
        return private_sub
    return gov_sub or private_sub


def licensing(value) -> str:
    """Strip the source's embedded field name. A blank stays blank — see the loader's note: a
    facility with no licensing status is overwhelmingly a Barangay Health Station, which is not a
    licensed facility type. Blank means "not stated", never "unlicensed"."""
    raw = text(value)
    if raw.startswith(LICENSING_PREFIX):
        raw = raw[len(LICENSING_PREFIX) :].strip()
    return raw


def clean(src_path: str, out_path: str) -> dict:
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"{src_path}: expected a sheet named {SHEET!r}, found {wb.sheetnames}")
    ws = wb[SHEET]

    rows = ws.iter_rows(values_only=True)
    header = [text(h) for h in next(rows)]
    if header != EXPECTED_HEADER:
        raise SystemExit(
            "Source header does not match the export this script was written against.\n"
            f"  expected {len(EXPECTED_HEADER)} columns, found {len(header)}\n"
            f"  first difference: "
            + next(
                (
                    f"col {i}: expected {e!r}, found {a!r}"
                    for i, (e, a) in enumerate(zip(EXPECTED_HEADER, header))
                    if e != a
                ),
                "column count only",
            )
        )

    problems: list[str] = []
    ownership_conflicts: list[str] = []
    out_rows: list[dict] = []
    types = Counter()
    regions = Counter()
    ownership = Counter()
    licensing_counts = Counter()
    no_barangay = 0
    stripped_licensing = 0

    for raw in rows:
        code = text(raw[SRC["facility_code"]])
        if not code:  # a trailing blank row, not a facility
            continue

        region_psgc = psgc(raw[SRC["region_psgc"]])
        province_psgc = psgc(raw[SRC["province_psgc"]])
        citymun_psgc = psgc(raw[SRC["citymun_psgc"]])
        barangay_psgc = psgc(raw[SRC["barangay_psgc"]])

        # Truncation to dim_geo widths is only lossless if the source's trailing zeros are where
        # the standard says. Check rather than assume — a code that fails this would silently
        # join to the wrong city.
        if region_psgc and not region_psgc.endswith("0" * 8):
            problems.append(f"{code}: region PSGC {region_psgc} is not region-shaped")
        if province_psgc and not province_psgc.endswith("0" * 5):
            problems.append(f"{code}: province PSGC {province_psgc} is not province-shaped")
        if citymun_psgc and not citymun_psgc.endswith("0" * 3):
            problems.append(f"{code}: city/municipality PSGC {citymun_psgc} is not citymun-shaped")
        if barangay_psgc and barangay_psgc[:7] != citymun_psgc[:7]:
            problems.append(
                f"{code}: barangay PSGC {barangay_psgc} is not inside city/municipality "
                f"{citymun_psgc}"
            )

        if not barangay_psgc:
            no_barangay += 1

        source_licensing = text(raw[SRC["licensing_status"]])
        if source_licensing.startswith(LICENSING_PREFIX):
            stripped_licensing += 1

        gov_sub = text(raw[SRC["ownership_sub_gov"]])
        private_sub = text(raw[SRC["ownership_sub_private"]])
        owner_major = text(raw[SRC["ownership_major"]])
        if gov_sub and private_sub:
            ownership_conflicts.append(
                f"{code} ({text(raw[SRC['facility_name']])}): {owner_major} — "
                f"government {gov_sub!r} + private {private_sub!r}"
            )

        row = {
            "facility_code": code,
            "facility_code_short": text(raw[SRC["facility_code_short"]]),
            "facility_name": text(raw[SRC["facility_name"]]),
            "facility_major_type": text(raw[SRC["facility_major_type"]]),
            "facility_type": text(raw[SRC["facility_type"]]),
            "ownership_major": owner_major,
            "ownership_sub": ownership_sub(gov_sub, private_sub, owner_major),
            "source_region_psgc": region_psgc,
            "source_region_name": text(raw[SRC["region_name"]]),
            "source_province_psgc": province_psgc,
            "source_province_name": text(raw[SRC["province_name"]]),
            "source_citymun_psgc": citymun_psgc,
            "source_citymun_name": text(raw[SRC["citymun_name"]]),
            "source_barangay_psgc": barangay_psgc,
            "source_barangay_name": text(raw[SRC["barangay_name"]]),
            "citymun_code": citymun_psgc[:7],
            "barangay_code": barangay_psgc,
            "service_capability": text(raw[SRC["service_capability"]]),
            "bed_capacity": bed_capacity(raw[SRC["bed_capacity"]], problems, code),
            "licensing_status": licensing(raw[SRC["licensing_status"]]),
            "license_validity_date": iso_date(raw[SRC["license_validity_date"]], problems, code),
        }

        if not row["facility_name"]:
            problems.append(f"{code}: blank facility name")
        if not row["citymun_code"]:
            problems.append(f"{code}: no city/municipality code — nothing to join on")

        out_rows.append(row)
        types[row["facility_type"]] += 1
        regions[row["source_region_name"]] += 1
        ownership[row["ownership_major"]] += 1
        licensing_counts[row["licensing_status"] or "(not stated)"] += 1

    codes = Counter(r["facility_code"] for r in out_rows)
    duplicates = [c for c, n in codes.items() if n > 1]
    if duplicates:
        problems.append(f"{len(duplicates)} duplicate facility codes, e.g. {duplicates[:5]}")

    if problems:
        shown = "\n".join(f"  - {p}" for p in problems[:50])
        more = f"\n  ... and {len(problems) - 50} more" if len(problems) > 50 else ""
        raise SystemExit(
            "Cleaning aborted — the source disagrees with what this script was written for:\n"
            + shown
            + more
        )

    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUT_COLUMNS)
        writer.writeheader()
        writer.writerows(out_rows)

    return {
        "rows": len(out_rows),
        "ownership_conflicts": ownership_conflicts,
        "types": types,
        "regions": regions,
        "ownership": ownership,
        "licensing": licensing_counts,
        "no_barangay": no_barangay,
        "stripped_licensing": stripped_licensing,
        "distinct_citymun": len({r["citymun_code"] for r in out_rows}),
        "distinct_barangay": len({r["barangay_code"] for r in out_rows if r["barangay_code"]}),
    }


def report(stats: dict, out_path: str) -> None:
    print(f"Wrote {stats['rows']:,} facilities to {out_path}")
    print()
    print(f"  Excluded columns ({len(EXCLUDED)}), each for a stated reason:")
    for col, why in EXCLUDED.items():
        print(f"    - {col}: {why}")
    print()
    print(f"  Licensing prefix stripped from {stats['stripped_licensing']:,} values")
    conflicts = stats["ownership_conflicts"]
    print(f"  Ownership sub-classification conflicts resolved by major classification: "
          f"{len(conflicts)}")
    for line in conflicts:
        print(f"    - {line}")
    print(f"  Facilities with no barangay code: {stats['no_barangay']:,} "
          f"(joined at city/municipality instead)")
    print(f"  Distinct city/municipalities: {stats['distinct_citymun']:,}")
    print(f"  Distinct barangays: {stats['distinct_barangay']:,}")
    print()
    print("  By region:")
    for name, n in stats["regions"].most_common():
        print(f"    {n:6,}  {name}")
    print()
    print("  By ownership:")
    for name, n in stats["ownership"].most_common():
        print(f"    {n:6,}  {name}")
    print()
    print("  By licensing status:")
    for name, n in stats["licensing"].most_common():
        print(f"    {n:6,}  {name}")
    print()
    print(f"  Facility types ({len(stats['types'])} distinct), top 15:")
    for name, n in stats["types"].most_common(15):
        print(f"    {n:6,}  {name}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--src", default=SRC_DEFAULT, help="Path to the NHFR export .xlsx")
    parser.add_argument("--out", default=OUT_DEFAULT, help="Path to write the cleaned CSV")
    args = parser.parse_args()

    stats = clean(args.src, args.out)
    report(stats, args.out)


if __name__ == "__main__":
    main()
