"""Turn the 2025 FHSIS annual workbooks into the committed long-format extracts the loader reads.

Source: `ingestion/data/fhsis_2025/`, pulled from DOH's public Drive archive by
`ingestion/fetch_fhsis.py`, which also writes the `_manifest.json` this script quotes for
provenance. Plan: `docs/FHSIS_2025_PLAN.md`, increment F1.

This mirrors `ingestion/clean_nhfr.py`: the cleaning step is separate from the load, the loader
never opens a workbook, and every action taken here is reported rather than applied silently. It
differs in one structural way — FHSIS is nine workbooks with nine different shapes, so the
per-sheet extraction spec below is data, and the code that walks it is generic.

**Decision 2, first and non-negotiably: FHSIS never supplies a BHW count.** The Demographics
`Health Workers` sheet publishes `Active Barangay Health Workers (BHW)` (col 36) and its ratio
(col 37). Both are dropped here, so they exist in no CSV and therefore in no table — the same
treatment `clean_nhfr.py` gives the contact columns, and for a stronger reason. This is not a
completeness gap to be filled later: the census (`bhw-2025`, with the StepZero quick-count as the
universe) is this site's BHW figure, and FHSIS's column is a tally of what LGUs filed through
their RHUs. The numbers in the file make the case by themselves — nationally 270,766, NCR 4,454,
and **Las Piñas 1**. Publishing that beside the census would undercut the census with a source
known to be under-reported. `fact_fhsis_workforce` additionally carries `check (cadre <> 'bhw')`
and the loader asserts zero such rows, so the rule is enforced three times over.

---

## What the workbooks actually turned out to be

Nine files were pulled; **six sheets across five workbooks are loaded**. Every exclusion below is
a property of the source found by reading it, and each corrects something `docs/FHSIS_2025_PLAN.md`
stated before the files were opened. The plan has been amended in the same pull request.

  - **There is no universal `Annual` sheet.** The plan said every workbook carries `Qtr1..Qtr4`
    plus `Annual`. Demographics carries `BGY & BHS` and `Health Workers` and no quarters at all;
    the three Envi files carry `Qtr1..Qtr4` and no `Annual`; TB carries `Annual` *and* three more
    tables that are part of the same cascade. So each sheet is named explicitly below.

  - **Envi is read from `Qtr4`, and that is the annual figure.** Water and sanitation are *stock*
    measures — households that have access as at the quarter end — not flows to be summed over
    four quarters. The year-end position is the Q4 sheet, whose title row reads "Philippines, 4th
    Quarter 2025". Summing the quarters would count the same household four times.

  - **The header is two merged rows only sometimes.** FIC/CIC, 4ANC, 8ANC and all four TB sheets
    are rows 4+5. Demographics `Health Workers` is rows 4+6 — row 5 is blank. Water is 4+5+6
    (three) and sanitation is 4+5+6+7 (four). A single-row read gives blanks, and so does a
    two-row read on half of these.

  - **`zod_nofml.xlsx` is 2024 data.** All four of its quarter sheets are titled "Philippines,
    Nth Quarter 2024" — the workbook sits in the 2025 folder but reports the year before. It is
    **not loaded**: `dim_dataset` registers this as the 2025 annual release with
    `as_of_date = 2025-12-31`, and a 2024 figure published under that row is a wrong citation,
    which Decision 1 exists to prevent. `zod` is therefore not in the indicator dictionary. When
    DOH publishes the 2025 ZOD table it loads through this script with no new code.

  - **2PNC and Demographics `BGY & BHS` have no PSGC column.** Both are ~146 rows keyed by area
    name — Tier 2 by the plan's own definition, needing the province-scoped matcher and fixups
    file that Decision 3 defers. Neither is loaded, so `pnc2` and the barangay/RHU/BHS counts are
    not in this increment. Deferred, not refused.

  - **8ANC's annual sheet covers Q3–Q4 2025, not the year.** Its title row says so:
    "Philippines, Q3-Q4 2025". It is loaded, because a half-year of a new indicator is still the
    published 2025 figure — but `ref_fhsis_indicator.numerator_def` carries the period, and no
    surface may compare an 8ANC rate with a 4ANC rate as though both covered twelve months.

  - **Not every rate is a percentage.** TB case notification and drug-resistant notification are
    published **per 100,000 population** — the national CNR is 473.06, which is 535,254 notified
    against a projected 113,146,216, not a 473% overshoot. TPT, TSR and every non-TB indicator are
    percentages. `ref_fhsis_indicator.unit` records which, and `over_100` is only ever set for a
    percentage indicator (see `is_over_100`). Marking a CNR as an overshoot would put a † on a
    normal figure and teach readers to ignore the marker.

  - **`DQC` is not a reliable column name.** 8ANC labels its check block `DQC`; 4ANC leaves the
    same block unlabelled with sub-headers like `C>=H`; FIC/CIC has two unlabelled trailing
    columns. So the rule here is the opposite of "drop the columns called DQC": **only declared
    columns are carried**, every other column is dropped, and the report names each dropped column
    with its header so an unnoticed indicator cannot hide among them.

## PSGC, which is where the real work is

The plan expected Tier 1 to join on `dim_geo.geo_code` "with no name-matching". It does not. About
70 rows per sheet carry a code that is wrong in a specific, repeatable way — leading zeros
stripped and the value right-padded back to width, so Ilocos Norte's Marcos (`0102813000`) is
printed as `128130000`. Resolution therefore runs in three stages, each counted separately in the
report so the mix is visible rather than averaged into a success rate:

  1. **Direct.** Normalise (Excel hands back floats: `1380100000.0`), zero-fill to ten, then
     truncate on the trailing-zero shape to `dim_geo`'s widths — region 2, province 5, citymun 7 —
     exactly as `clean_nhfr.py` truncates. This resolves 1,671–1,743 rows of every sheet.
  2. **Shift repair.** One leading zero lost and one zero appended: drop the trailing zero,
     re-fill, retry. 0–15 rows per sheet. Applied *only* after the direct read fails, which
     matters — NCR's `1300000000` shift-repairs to a valid Region I code, so trying the repair
     first would silently move NCR's figures to Ilocos.
  3. **Parent-scoped name match.** For the rest (57 rows per affected sheet, all in Region I, plus
     three province rows), match the printed area name among the children of the *already
     resolved* parent — the last region or province row above it — using `ingest_population.py`'s
     `variants()`, the normaliser this repo already trusts for PSA name reconciliation. Scoped to
     one parent, so "San Nicolas" cannot match the San Nicolas in a different province. This is
     name matching, which Decision 3 said Tier 1 would not need; it is confined to a repair path
     that fires on 3% of rows, and every one of them is listed in the report.

After all three, **0 rows fail to resolve** in any loaded sheet, which is what the F1 check
requires. A row that resolved by stage 2 or 3 still carries its `source_psgc` and
`source_area_name` into the fact table, so the repair is visible in the data and not just here.

## What is stored, per Decision 4

Numerator, denominator and the rate **as the source printed it**, per geo per indicator per
breakdown, at every grain the source publishes (national, region, province/HUC, citymun). Nothing
is recomputed and nothing is capped: a rate above 100 is stored with `over_100 = true` because the
denominator is present and the overshoot is explainable, which is more honest than a ceiling.
Rollups are not derived from leaves — the source publishes its own subtotals, and the citymun
leaves are known to be incomplete. Instead the leaf-vs-published-parent residual is computed per
indicator and written to the report and to `ref_fhsis_reconciliation`.

Usage:
    python ingestion/clean_fhsis.py --dim-geo-csv ingestion/data/dim_geo.csv
    python ingestion/clean_fhsis.py --database-url "$SUPABASE_DB_URL"
"""

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

# The repo's existing name normaliser, written for PSA population reconciliation. Reused rather
# than reimplemented so a spelling that was already reconciled once (BALIUAG/BALIWAG and the rest
# of SPELLING_FIXUPS) stays reconciled here, in one place.
from ingest_population import variants  # noqa: E402

SRC_DIR_DEFAULT = "ingestion/data/fhsis_2025"
OUT_DIR_DEFAULT = "ingestion/data"
REPORT_DEFAULT = "docs/FHSIS_2025_CLEANING_REPORT.md"
MANIFEST_NAME = "_manifest.json"

# The machine-readable twin of the report. Written by the same run, from the same counters, and
# read by ingest_fhsis.py — which is what makes the plan's "row counts match the cleaning report"
# and "over_100 count per indicator matches the report" real checks rather than two lists of
# numbers that drift apart. The prose report is for a person; this is for the loader.
SUMMARY_DEFAULT = "ingestion/data/fhsis_2025_cleaning_summary.json"

# The headline figures the plan requires this load to reproduce. Checked here as well as in the
# loader: a cleaning run that has already lost them should say so before anything is written.
NATIONAL_CHECKS = {
    ("fic", "total", "numerator"): 1_560_924,
    ("fic", "total", "denominator"): 2_392_392,
}
NATIONAL_POPULATION_2025 = 113_146_216
NATIONAL_HOUSEHOLDS_2025 = 27_387_195

# Percent indicators may exceed 100 and are marked when they do. Rate indicators are per 100,000
# population, where a value above 100 is ordinary — see the module docstring.
UNIT_PERCENT = "percent (0-100, may exceed)"
UNIT_PER_100K = "cases per 100,000 population"

# --------------------------------------------------------------------------- #
# The indicator dictionary.                                                     #
#                                                                               #
# This is the source of truth for ref_fhsis_indicator: the migration seeds       #
# exactly these rows, and the loader refuses any indicator_key not listed here.  #
# numerator_def / denominator_def are the workbook's own header text, so a       #
# reader can find the column the figure came from.                              #
# --------------------------------------------------------------------------- #
INDICATORS = {
    # -- immunization ------------------------------------------------------- #
    "fic": {
        "program_area": "immunization",
        "label": "Fully immunised children (FIC)",
        "numerator_def": "Fully Immunized Children, annual 2025",
        "denominator_def": "Projected Population (0-12 months old) (Previous cohort)",
        "unit": UNIT_PERCENT,
        "source_workbook": "9 10 FIC _ CIC_MunCity EB 2025_FInal Annual 2025_NF.xlsx",
        "uuc_criterion_d": True,
    },
    "cic": {
        "program_area": "immunization",
        "label": "Completely immunised children (CIC)",
        "numerator_def": "Completely Immunized Children, annual 2025",
        "denominator_def": "Projected population 0-12 months (previous cohort) minus previous FIC",
        "unit": UNIT_PERCENT,
        "source_workbook": "9 10 FIC _ CIC_MunCity EB 2025_FInal Annual 2025_NF.xlsx",
        "uuc_criterion_d": False,
    },
    # -- maternal ----------------------------------------------------------- #
    "anc4": {
        "program_area": "maternal",
        "label": "Women who completed at least 4 antenatal care visits",
        "numerator_def": "Total No. of women who delivered and completed at least 4ANC",
        "denominator_def": "Total Deliveries",
        "unit": UNIT_PERCENT,
        "source_workbook": "1. 4ANC_2025_EB_1_nofml.xlsx",
        "uuc_criterion_d": True,
    },
    "anc8": {
        "program_area": "maternal",
        # The period is in the label, not only in a note, because this figure is not
        # year-comparable with anc4 and every surface that renders it must say so.
        "label": "Women who completed at least 8 antenatal care visits (Q3-Q4 2025 only)",
        "numerator_def": (
            "Total No. of women who delivered and completed at least 8ANC (d+e). The source "
            "sheet is titled 'Philippines, Q3-Q4 2025': it covers the second half of 2025 only, "
            "because 8ANC was introduced mid-year. Not comparable with anc4's full year."
        ),
        "denominator_def": (
            "Total No. of women who delivered and were tracked during pregnancy (a+b)-c, "
            "Q3-Q4 2025"
        ),
        "unit": UNIT_PERCENT,
        "source_workbook": "1. 8ANC_2025_PH_nofml.xlsx",
        "uuc_criterion_d": True,
    },
    # -- envi --------------------------------------------------------------- #
    "water_basic": {
        "program_area": "envi",
        "label": "Households with access to basic safe water supply",
        "numerator_def": "Households with Access to Basic Safe Water Supply, Total (Levels I-III)",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "water_nofml.xlsx",
        "uuc_criterion_d": True,
    },
    "water_level1": {
        "program_area": "envi",
        "label": "Households with Level I water supply (point source)",
        "numerator_def": "Households with Access to Basic Safe Water Supply, Level 1",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "water_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "water_level2": {
        "program_area": "envi",
        "label": "Households with Level II water supply (communal faucet)",
        "numerator_def": "Households with Access to Basic Safe Water Supply, Level 2",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "water_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "water_level3": {
        "program_area": "envi",
        "label": "Households with Level III water supply (waterworks)",
        "numerator_def": "Households with Access to Basic Safe Water Supply, Level 3",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "water_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "water_safely_managed": {
        "program_area": "envi",
        "label": "Households using safely managed drinking water services",
        "numerator_def": "Households using Safely Managed Drinking Water Services",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "water_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "sanitation_basic": {
        "program_area": "envi",
        "label": "Households with a basic sanitation facility",
        "numerator_def": "Households with Basic Sanitation Facility, Total",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "sanitation_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "sanitation_septic": {
        "program_area": "envi",
        "label": "Households with a pour/flush toilet to a septic tank",
        "numerator_def": "Households with Basic Sanitation Facility, Pour/Flush Toilet - Septic Tank",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "sanitation_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "sanitation_sewer": {
        "program_area": "envi",
        "label": "Households with a pour/flush toilet to a community sewer",
        "numerator_def": (
            "Households with Basic Sanitation Facility, Pour/Flush Toilet - Community "
            "Sewer/Sewerage System"
        ),
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "sanitation_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "sanitation_vip": {
        "program_area": "envi",
        "label": "Households with a ventilated improved pit latrine",
        "numerator_def": (
            "Households with Basic Sanitation Facility, Ventilated Improved Pit (VIP) Latrine"
        ),
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "sanitation_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    "sanitation_safely_managed": {
        "program_area": "envi",
        "label": "Households using safely managed sanitation services",
        "numerator_def": "Households using Safely Managed Sanitation Service",
        "denominator_def": "Projected Number of Households",
        "unit": UNIT_PERCENT,
        "source_workbook": "sanitation_nofml.xlsx",
        "uuc_criterion_d": False,
    },
    # -- tb ----------------------------------------------------------------- #
    "tb_notified": {
        "program_area": "tb",
        "label": "TB case notification rate, all forms",
        "numerator_def": "No. of notified TB cases, all forms",
        "denominator_def": "Projected Population (All Ages)",
        # Per 100,000, not percent. The national figure is 473.06 — see the module docstring.
        "unit": UNIT_PER_100K,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
    "tb_dr_notified": {
        "program_area": "tb",
        "label": "Drug-resistant TB notification rate (RR/MDR-TB)",
        "numerator_def": (
            "No. registered bacteriologically confirmed drug-resistant TB (RR/MDR-TB Cases)"
        ),
        "denominator_def": "Projected Population (All Ages)",
        "unit": UNIT_PER_100K,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
    "tb_presumptive_tested": {
        "program_area": "tb",
        "label": "Presumptive TB tested with a bacteriologic test",
        "numerator_def": "No. of presumptive Tuberculosis tested with bacteriologic test",
        "denominator_def": "Projected Population (All Ages)",
        "unit": UNIT_PERCENT,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
    "tb_tpt": {
        "program_area": "tb",
        "label": "TB preventive treatment (TPT) coverage for TB contacts",
        "numerator_def": "No. of TB contacts given TB Preventive Treatment (All Ages)",
        "denominator_def": "Eligible Population (Notified TB cases x 4 contacts x 70%)",
        "unit": UNIT_PERCENT,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
    "tb_tsr_dstb": {
        "program_area": "tb",
        "label": "Treatment success rate, drug-susceptible TB (all forms)",
        "numerator_def": (
            "No. of TB all forms that are cured and completely treated (All ages)"
        ),
        # The denominator is the case count, not a population — the one indicator here whose
        # denominator is not a projection, which is why it is spelled out.
        "denominator_def": "No. of TB all forms (all ages) — cases registered, not a population",
        "unit": UNIT_PERCENT,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
    "tb_tsr_mdrtb": {
        "program_area": "tb",
        "label": "Treatment success rate, drug-resistant TB (RR/MDR-TB)",
        "numerator_def": (
            "No. of registered bacteriology confirmed drug resistant TB Cases (RR/MDR-TB) cured "
            "and completed treatment (All Ages)"
        ),
        "denominator_def": (
            "No. of registered bacteriology confirmed drug resistant TB Cases (RR/MDR-TB) "
            "(all ages) — cases registered, not a population"
        ),
        "unit": UNIT_PERCENT,
        "source_workbook": "FINAL Tuberculosis_EB_2025.xlsx",
        "uuc_criterion_d": False,
    },
}

# The breakdown vocabulary, as the source turned out to publish it. The plan named
# total|male|female|10-14|15-19|20-49; TB's treatment and preventive-treatment tables add the two
# paediatric/adult bands below, so they are part of the vocabulary rather than dropped.
BREAKDOWNS = ("total", "male", "female", "10-14", "15-19", "20-49", "0-14", "15+")

# The eight cadres, source label -> the key fact_fhsis_workforce stores. 'Nutrtionists' is the
# workbook's own spelling; matched as printed rather than corrected, so a future re-export that
# fixes the typo fails the header check loudly instead of silently dropping the cadre.
CADRES = {
    "Doctors": "physician",
    "Nurses": "nurse",
    "Midwives": "midwife",
    "Dentists": "dentist",
    "Medical Technologists": "med_tech",
    "Nutrtionists": "nutritionist",
    "Sanitary Engineers": "sanitary_engineer",
    "Sanitary Inspectors": "sanitary_inspector",
}


def _immunization_measures() -> dict:
    return {
        ("fic", "total"): {"numerator": 5, "denominator": 2, "rate_pct": 6},
        ("fic", "male"): {"numerator": 3},
        ("fic", "female"): {"numerator": 4},
        ("cic", "total"): {"numerator": 10, "denominator": 7, "rate_pct": 11},
        ("cic", "male"): {"numerator": 8},
        ("cic", "female"): {"numerator": 9},
    }


def _anc4_measures() -> dict:
    return {
        ("anc4", "10-14"): {"numerator": 7, "denominator": 2, "rate_pct": 8},
        ("anc4", "15-19"): {"numerator": 9, "denominator": 3, "rate_pct": 10},
        ("anc4", "20-49"): {"numerator": 11, "denominator": 4, "rate_pct": 12},
        ("anc4", "total"): {"numerator": 13, "denominator": 5, "rate_pct": 14},
    }


def _anc8_measures() -> dict:
    return {
        ("anc8", "10-14"): {"numerator": 27, "denominator": 14, "rate_pct": 28},
        ("anc8", "15-19"): {"numerator": 29, "denominator": 15, "rate_pct": 30},
        ("anc8", "20-49"): {"numerator": 31, "denominator": 16, "rate_pct": 32},
        ("anc8", "total"): {"numerator": 33, "denominator": 17, "rate_pct": 34},
    }


def _water_measures() -> dict:
    return {
        ("water_level1", "total"): {"numerator": 3, "denominator": 2, "rate_pct": 4},
        ("water_level2", "total"): {"numerator": 5, "denominator": 2, "rate_pct": 6},
        ("water_level3", "total"): {"numerator": 7, "denominator": 2, "rate_pct": 8},
        ("water_basic", "total"): {"numerator": 9, "denominator": 2, "rate_pct": 10},
        ("water_safely_managed", "total"): {"numerator": 11, "denominator": 2, "rate_pct": 12},
    }


def _sanitation_measures() -> dict:
    return {
        ("sanitation_septic", "total"): {"numerator": 3, "denominator": 2, "rate_pct": 4},
        ("sanitation_sewer", "total"): {"numerator": 5, "denominator": 2, "rate_pct": 6},
        ("sanitation_vip", "total"): {"numerator": 7, "denominator": 2, "rate_pct": 8},
        ("sanitation_basic", "total"): {"numerator": 9, "denominator": 2, "rate_pct": 10},
        ("sanitation_safely_managed", "total"): {
            "numerator": 11,
            "denominator": 2,
            "rate_pct": 12,
        },
    }


def _tb_annual_measures() -> dict:
    return {
        ("tb_notified", "total"): {"numerator": 5, "denominator": 2, "rate_pct": 6},
        ("tb_notified", "male"): {"numerator": 3},
        ("tb_notified", "female"): {"numerator": 4},
        ("tb_dr_notified", "total"): {"numerator": 9, "denominator": 2, "rate_pct": 10},
        ("tb_dr_notified", "male"): {"numerator": 7},
        ("tb_dr_notified", "female"): {"numerator": 8},
        ("tb_presumptive_tested", "total"): {"numerator": 13, "denominator": 2, "rate_pct": 14},
        ("tb_presumptive_tested", "male"): {"numerator": 11},
        ("tb_presumptive_tested", "female"): {"numerator": 12},
    }


def _tb_tpt_measures() -> dict:
    return {
        ("tb_tpt", "total"): {"numerator": 11, "denominator": 8, "rate_pct": 12},
        ("tb_tpt", "male"): {"numerator": 9},
        ("tb_tpt", "female"): {"numerator": 10},
        ("tb_tpt", "0-14"): {"numerator": 4},
        ("tb_tpt", "15+"): {"numerator": 7},
    }


def _tb_tsr_measures(key: str) -> dict:
    """TSR-DSTB and TSR-MDRTB share a column layout exactly; only the case definition differs."""
    return {
        (key, "total"): {"numerator": 22, "denominator": 19, "rate_pct": 23},
        (key, "male"): {"numerator": 20},
        (key, "female"): {"numerator": 21},
        (key, "0-14"): {"numerator": 6, "denominator": 3, "rate_pct": 7},
        (key, "15+"): {"numerator": 12, "denominator": 9, "rate_pct": 13},
    }


# --------------------------------------------------------------------------- #
# The per-sheet extraction specs.                                              #
#                                                                              #
# `expect` is the load-bearing part: the merged header text this script was     #
# written against, per column it reads. A re-export that reshapes a table fails  #
# here naming the column, rather than silently reading the wrong one — the same  #
# guarantee clean_nhfr.py's EXPECTED_HEADER gives, made per column because these #
# sheets are too wide for a whole-row comparison to be readable.                 #
# --------------------------------------------------------------------------- #
SHEETS = [
    {
        "workbook": "demographics",
        "area": "demographics",
        "file": "Demographic_2025_EB_Final.xlsx",
        "sheet": "Health Workers",
        "header_rows": [4, 6],
        "data_start": 7,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "workforce",
        "period": "annual 2025",
        "expect": {
            0: "PSGC",
            1: "Area",
            2: "Population 2025",
            3: "Doctors || LGU Hired",
            35: "Number of Household Estimates",
        },
        "population_col": 2,
        "households_col": 35,
        # Cadre blocks start at column 3 and repeat every four columns
        # (LGU Hired, DOH Hired, Total, Ratio) in CADRES order.
        "cadre_first_col": 3,
        "cadre_stride": 4,
        "dropped": {
            36: "Active Barangay Health Workers (BHW), LGU Hired — plan Decision 2, never loaded",
            37: "Active Barangay Health Workers (BHW), Ratio — plan Decision 2, never loaded",
        },
    },
    {
        "workbook": "immunization_fic_cic",
        "area": "immunization",
        "file": "9 10 FIC _ CIC_MunCity EB 2025_FInal Annual 2025_NF.xlsx",
        "sheet": "Annual",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            2: "Projected Population (0-12 months old) (Previous cohort)",
            5: "FIC || Total",
            6: "FIC || %",
            7: "Projected population 0-12 months (previous cohort) minus previous FIC",
            10: "CIC || Total",
        },
        "measures": _immunization_measures(),
    },
    {
        "workbook": "maternal_anc4",
        "area": "maternal",
        "file": "1. 4ANC_2025_EB_1_nofml.xlsx",
        "sheet": "Annual",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            2: "Total Deliveries || 10-14 years old",
            5: "Total Deliveries || Total",
            13: "Total No. of women who delivered and completed at least 4ANC || Total",
            14: "Total No. of women who delivered and completed at least 4ANC || %",
        },
        "measures": _anc4_measures(),
    },
    {
        "workbook": "maternal_anc8",
        "area": "maternal",
        "file": "1. 8ANC_2025_PH_nofml.xlsx",
        "sheet": "Annual",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        # Recorded, and carried into the report, because the sheet is not a full year.
        "period": "Q3-Q4 2025 (the source sheet's own title)",
        "expect": {
            0: "PSGC 10",
            17: "Total No. of women who delivered and were tracked during pregnancy (a+b)-c "
                "|| Total",
            33: "Total No. of women who delivered and completed at least 8ANC (d+e) || Total",
            34: "Total No. of women who delivered and completed at least 8ANC (d+e) || %",
        },
        "measures": _anc8_measures(),
    },
    {
        "workbook": "envi_water",
        "area": "envi",
        "file": "water_nofml.xlsx",
        "sheet": "Qtr4",
        "header_rows": [4, 5, 6],
        "data_start": 7,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "4th quarter 2025 (a year-end stock, used as the annual figure)",
        "expect": {
            0: "PSGC",
            2: "Projected Number of Households",
            3: "Households with Access to Basic Safe Water Supply || Level 1 || Number",
            9: "Households with Access to Basic Safe Water Supply || Total",
            10: "Households with Access to Basic Safe Water Supply || %",
            11: "Households using Safely Managed Drinking Water Services || Number",
        },
        "measures": _water_measures(),
    },
    {
        "workbook": "envi_sanitation",
        "area": "envi",
        "file": "sanitation_nofml.xlsx",
        "sheet": "Qtr4",
        "header_rows": [4, 5, 6, 7],
        "data_start": 8,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "4th quarter 2025 (a year-end stock, used as the annual figure)",
        "expect": {
            0: "PSGC",
            2: "Projected Number of Households",
            9: "Households with Basic Sanitation Facility || Pour/Flush Toilet || Total",
            10: "Households with Basic Sanitation Facility || Pour/Flush Toilet || %",
            11: "Households with Basic Sanitation Facility || Households using Safely Managed "
                "Sanitation Service || Number",
        },
        "measures": _sanitation_measures(),
    },
    {
        "workbook": "tb",
        "area": "tb",
        "file": "FINAL Tuberculosis_EB_2025.xlsx",
        "sheet": "Annual",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            2: "Projected Population (All Ages)",
            5: "No. of notified TB cases, all forms || Total",
            6: "No. of notified TB cases, all forms || %",
            9: "No. registered bacteriologically confirmed drug-resistant TB (RR/MDR-TB Cases) "
               "|| Total",
        },
        "measures": _tb_annual_measures(),
    },
    {
        "workbook": "tb",
        "area": "tb",
        "file": "FINAL Tuberculosis_EB_2025.xlsx",
        "sheet": "TB-TPT",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            4: "No. of TB contacts given TB Preventive Treatment (0-14 years old) || Total",
            7: "No. of TB contacts given TB Preventive Treatment (15 yrs old and above) || Total",
            8: "Eligible Population (Notified TB cases x 4 contacts x 70%)",
            11: "No. of TB contacts given TB Preventive Treatment (All Ages) || Total",
        },
        "measures": _tb_tpt_measures(),
    },
    {
        "workbook": "tb",
        "area": "tb",
        "file": "FINAL Tuberculosis_EB_2025.xlsx",
        "sheet": "TSR-DSTB",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            3: "No. of TB all forms (0-14)",
            9: "No. of TB all forms (15 years old and above)",
            19: "No. of TB all forms (all ages)",
            22: "No. of TB all forms that are cured and completely treated (All ages) || Total",
        },
        "measures": _tb_tsr_measures("tb_tsr_dstb"),
    },
    {
        "workbook": "tb",
        "area": "tb",
        "file": "FINAL Tuberculosis_EB_2025.xlsx",
        "sheet": "TSR-MDRTB",
        "header_rows": [4, 5],
        "data_start": 6,
        "psgc_col": 0,
        "name_col": 1,
        "kind": "indicator",
        "period": "annual 2025",
        "expect": {
            0: "PSGC 10",
            3: "No. of registered bacteriology confirmed drug resistant TB Cases (RR/MDR-TB) "
               "(0-14)",
            19: "No. of registered bacteriology confirmed drug resistant TB Cases (RR/MDR-TB) "
                "(all ages)",
            22: "No. of registered bacteriology confirmed drug resistant TB Cases (RR/MDR-TB) "
                "cured and completed treatment (All Ages) || Total",
        },
        "measures": _tb_tsr_measures("tb_tsr_mdrtb"),
    },
]

# Workbooks fetched but deliberately not loaded, each with the reason. Named rather than merely
# omitted, on clean_nhfr.py's EXCLUDED precedent: an exclusion a reader can audit is a decision,
# an exclusion they cannot see is an oversight.
EXCLUDED_WORKBOOKS = {
    "envi_zod": (
        "every quarter sheet is titled 'Philippines, Nth Quarter 2024' — the workbook is 2024 "
        "data sitting in the 2025 folder. Loading it under a dataset row that says 2025 would "
        "be a wrong citation (plan Decision 1)."
    ),
    "maternal_pnc2": (
        "no PSGC column: 146 rows keyed by area name only, i.e. Tier 2, which needs the "
        "province-scoped matcher and fixups file that plan Decision 3 defers."
    ),
}
EXCLUDED_SHEETS = {
    "Demographic_2025_EB_Final.xlsx / BGY & BHS": (
        "no PSGC column: 146 rows keyed by area name only. The barangay / RHU / BHS counts are "
        "Tier 2 and deferred with the rest of it (plan Decision 3)."
    ),
}

OUT_COLUMNS_INDICATOR = [
    "indicator_key",
    "breakdown",
    "geo_code",
    "geo_level",
    "numerator",
    "denominator",
    "rate_pct",
    "over_100",
    "source_psgc",
    "source_area_name",
    "geo_resolution",
]

OUT_COLUMNS_WORKFORCE = [
    "cadre",
    "geo_code",
    "geo_level",
    "lgu_hired",
    "doh_hired",
    "total",
    "population_2025",
    "households_2025",
    "source_psgc",
    "source_area_name",
    "geo_resolution",
]


# --------------------------------------------------------------------------- #
# Cell reading                                                                 #
# --------------------------------------------------------------------------- #
def text(value) -> str:
    """Cell -> trimmed single-spaced string. Newlines inside merged headers become spaces."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def raw_psgc(value) -> str:
    """PSGC cell -> the digits the source printed, before any repair.

    openpyxl hands the same column back as a float on some rows (`1380100000.0`) and as text on
    others; both are the same code and the trailing `.0` is a storage artefact, not a value.
    """
    s = text(value)
    return s[:-2] if s.endswith(".0") else s


def number(value, problems: list, where: str):
    """Data cell -> int/float, or None for 'not reported'.

    The workbooks' own legend defines the two non-numeric values: an asterisk is
    "incomplete data/no data submitted" and a zero is "zero data/zero cases". So `*` becomes None
    — *not reported* — and 0 stays 0. Decision 7 turns on exactly this distinction: a city with a
    zero delivered zero and the page says so; a city with no figure is not reporting and the page
    says that instead. Collapsing them would invent a zero.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    s = text(value)
    if s == "" or set(s) <= {"*"} or s in {"-", "--", "n/a", "N/A"}:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        problems.append(f"{where}: non-numeric value {s!r}")
        return None


def as_int(value):
    """A count column -> int. The source stores counts as floats, and a handful of Envi cells
    carry a fractional household estimate (`5297238.1826`) because the figure is a projection
    apportioned across levels. Rounded rather than truncated, and only ever for a count."""
    if value is None:
        return None
    return int(round(float(value)))


def merged_headers(worksheet, header_rows: list[int], width: int) -> list[str]:
    """The header rows folded into one label per column, honouring the real merge ranges.

    A merged cell holds its value only in its top-left cell, so a naive forward-fill is the
    obvious way to rebuild the label — and it is wrong at the right-hand end of a row, where it
    carries a group heading on past the columns it actually spans and stamps it onto the DQC
    block that follows. So the spans come from `worksheet.merged_cells.ranges` instead, and a
    column outside every span keeps its own blank.

    This is the plan's "build the column map from both rows" generalised to the two, three and
    four header rows the workbooks actually use. Parts are de-duplicated down the column, because
    a vertically merged heading otherwise repeats itself once per row it spans.
    """
    grid: dict[tuple[int, int], str] = {}
    for row in header_rows:
        for col in range(width):
            grid[(row, col)] = text(worksheet.cell(row=row + 1, column=col + 1).value)
    for span in worksheet.merged_cells.ranges:
        value = text(worksheet.cell(row=span.min_row, column=span.min_col).value)
        for row in range(span.min_row - 1, span.max_row):
            if row not in header_rows:
                continue
            for col in range(span.min_col - 1, min(span.max_col, width)):
                grid[(row, col)] = value

    labels = []
    for col in range(width):
        parts, seen = [], set()
        for row in header_rows:
            part = grid.get((row, col), "")
            if part and part not in seen:
                seen.add(part)
                parts.append(part)
        labels.append(" || ".join(parts))
    return labels


# --------------------------------------------------------------------------- #
# Geography                                                                    #
# --------------------------------------------------------------------------- #
def load_dim_geo_csv(path: str) -> dict:
    with open(path, newline="", encoding="utf-8") as fh:
        return {
            r["geo_code"]: (r["geo_level"], r["geo_name"]) for r in csv.DictReader(fh)
        }


def load_dim_geo_from_db(database_url: str) -> dict:
    import psycopg2

    conn = psycopg2.connect(database_url)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "select geo_code, geo_level::text, geo_name from dim_geo "
                "where geo_level in ('national','region','province','citymun')"
            )
            return {code: (level, name) for code, level, name in cur.fetchall()}
    finally:
        conn.close()


def truncate_to_dim_geo(code: str) -> tuple[str, str] | None:
    """A 10-digit PSGC -> (dim_geo code, level), by the trailing-zero shape.

    dim_geo stores region at 2 characters, province at 5 and city/municipality at 7, so the
    truncation is the same one clean_nhfr.py performs; the shape is *checked* rather than
    assumed, because a code that fails it would otherwise join to the wrong place.
    """
    if not re.fullmatch(r"\d{10}", code):
        return None
    if code.endswith("0" * 8):
        return code[:2], "region"
    if code.endswith("0" * 5):
        return code[:5], "province"
    if code.endswith("0" * 3):
        return code[:7], "citymun"
    return None


class GeoResolver:
    """Three-stage PSGC resolution. See the module docstring for why each stage exists."""

    def __init__(self, dim_geo: dict):
        self.dim_geo = dim_geo
        # name -> code, indexed by (parent code, level), so a name match is always scoped to one
        # parent and "San Nicolas" cannot cross a province boundary.
        self.children: dict[tuple[str, str], dict[str, str]] = {}
        for code, (level, name) in dim_geo.items():
            if level == "citymun":
                parent = code[:5]
            elif level == "province":
                parent = code[:2]
            elif level == "region":
                parent = "PH"
            else:
                continue
            index = self.children.setdefault((parent, level), {})
            for variant in variants(name):
                index.setdefault(variant, code)
        self.reset()

    def reset(self) -> None:
        """Called per sheet: the parent scope is positional, so it must not leak across sheets."""
        self.current = {"region": None, "province": None}
        self.stats = Counter()
        self.repairs: list[str] = []
        self.overrides: list[str] = []
        self.naming_differences: set[str] = set()

    def _name_agrees(self, area_name: str, code: str) -> bool:
        """Does the printed area name describe the geography the code resolved to?"""
        target = set(variants(self.dim_geo[code][1]))
        return any(v in target for v in variants(area_name))

    def _by_name(self, area_name: str):
        """Parent-scoped name lookup -> (code, level) or None. Reads the *current* scope, so it
        must be called before this row updates it."""
        for level, parent in (
            ("citymun", self.current["province"]),
            ("province", self.current["region"]),
        ):
            if not parent:
                continue
            index = self.children.get((parent, level), {})
            for variant in variants(area_name):
                if variant in index:
                    return index[variant], level
        return None

    def _remember(self, code: str, level: str) -> None:
        if level == "region":
            self.current["region"] = code
            # A new region invalidates the province below it; without this a name match under the
            # last province of the previous region could win.
            self.current["province"] = None
        elif level == "province":
            self.current["province"] = code

    def _accept(self, code: str, level: str, how: str, source_psgc: str, area_name: str):
        """Take a code the PSGC resolved to — unless the printed name says it is the wrong place.

        The source contains at least one block whose codes are swapped: in every sheet that has
        it, the province row "Surigao del Norte" carries `1606701000` (Alegria's code) and the
        row "Alegria" carries `1606700000` (the province's). Honouring the code alone files a
        province's 3,118 antenatal visits under a municipality that had 9 of them, and the error
        is invisible — both codes resolve perfectly.

        So the name is used as a **check on** the code, never as the join key (Decision 3's
        no-name-matching rule is about the join, and this is not one). The override is deliberately
        narrow and fires only when the printed name resolves, inside the current parent scope, to
        a *different* geography. Across all ten sheets that is one row pair. Thirty-six other rows
        disagree in wording only — "Region 1" against dim_geo's "REGION I (ILOCOS REGION)",
        "Davao del Oro" against "DAVAO DE ORO", the BARMM Special Geographic Area's cluster names
        against their barangay-cluster names — and in every one of those the name resolves to
        nothing else, so the code stands. Both outcomes are counted and listed in the report,
        which is BUILD_PLAN P15's discipline: reconcile by a stated rule, log the exceptions.
        """
        if self._name_agrees(area_name, code):
            self._remember(code, level)
            self.stats[how] += 1
            return code, level, how

        alternative = self._by_name(area_name)
        if alternative and alternative[0] != code:
            alt_code, alt_level = alternative
            self._remember(alt_code, alt_level)
            self.stats["name_override"] += 1
            self.overrides.append(
                f"{source_psgc} printed as {area_name!r}: the code resolves to {code} "
                f"({self.dim_geo[code][1]}) but the name resolves to {alt_code} "
                f"({self.dim_geo[alt_code][1]}) inside the same parent — the name is honoured"
            )
            return alt_code, alt_level, "name_override"

        self._remember(code, level)
        self.stats[how] += 1
        self.naming_differences.add(
            f"{source_psgc} printed as {area_name!r} -> {code} "
            f"({self.dim_geo[code][1]}) — wording differs, code honoured"
        )
        return code, level, how

    def resolve(self, source_psgc: str, area_name: str) -> tuple[str, str, str] | None:
        """-> (dim_geo code, level, how) or None."""
        digits = source_psgc
        if digits and digits.isdigit() and len(digits) < 10:
            digits = digits.zfill(10)

        direct = truncate_to_dim_geo(digits) if digits else None
        if direct and direct[0] in self.dim_geo:
            return self._accept(direct[0], direct[1], "direct", source_psgc, area_name)

        # Stage 2 runs only after stage 1 has failed. NCR's 1300000000 shift-repairs to a valid
        # Region I code, so the order is what keeps NCR's figures in NCR.
        if digits and digits.isdigit() and digits.endswith("0"):
            shifted = truncate_to_dim_geo(digits[:-1].zfill(10))
            if shifted and shifted[0] in self.dim_geo:
                self.repairs.append(
                    f"shift: {source_psgc} -> {shifted[0]} ({self.dim_geo[shifted[0]][1]}) "
                    f"[source name: {area_name}]"
                )
                return self._accept(
                    shifted[0], shifted[1], "shift_repair", source_psgc, area_name
                )

        found = self._by_name(area_name)
        if found:
            code, level = found
            self._remember(code, level)
            self.stats["name_scoped"] += 1
            self.repairs.append(
                f"name: {source_psgc or '(blank)'} + {area_name!r} -> {code} "
                f"({self.dim_geo[code][1]})"
            )
            return code, level, "name_scoped"

        self.stats["unresolved"] += 1
        return None


# --------------------------------------------------------------------------- #
# Extraction                                                                   #
# --------------------------------------------------------------------------- #
def is_over_100(indicator_key: str, rate) -> bool:
    """Decision 4's marker, but only for the indicators where it means anything.

    A percentage above 100 is a real overshoot — more services recorded than the projected
    eligible population — and gets the † the plan requires. A TB notification rate is per 100,000
    and is *normally* in the hundreds; flagging 473.06 as an overshoot would put a warning on an
    ordinary figure, and a marker that fires on ordinary figures stops being read.
    """
    if rate is None:
        return False
    return INDICATORS[indicator_key]["unit"] == UNIT_PERCENT and float(rate) > 100


def is_data_row(row: tuple) -> bool:
    return any(c is not None and str(c).strip() for c in row)


def looks_like_note(psgc: str, name: str) -> bool:
    """Trailing legend/source rows, which sit in the PSGC column, not the name column.

    Every sheet ends with a block like 'Source: DOH-Field Health Services Information System
    (FHSIS)', 'Legend:', 'Asterisk (*) - incomplete data/no data submitted'. They are recognised
    and skipped rather than reported as unresolvable geographies.

    Only a *non-empty, non-numeric* PSGC marks a note. The national row has a blank PSGC and the
    name 'PHILIPPINES', so testing the name here would throw the national row away — which is the
    one row the load-blocking FIC check reads.
    """
    del name  # deliberately unused: see the docstring
    return bool(psgc) and not psgc.replace(".", "").isdigit()


def extract_sheet(spec: dict, src_dir: Path, resolver: GeoResolver, problems: list) -> dict:
    path = src_dir / spec["file"]
    # Not read_only: merged_cells.ranges is what makes the header labels exact, and a read-only
    # worksheet does not expose it. These sheets are ~1,750 rows, so the cost is a non-issue.
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        if spec["sheet"] not in workbook.sheetnames:
            raise SystemExit(
                f"{spec['file']}: expected a sheet named {spec['sheet']!r}, found "
                f"{workbook.sheetnames}"
            )
        worksheet = workbook[spec["sheet"]]
        rows = list(worksheet.iter_rows(values_only=True))
        where = f"{spec['file']} [{spec['sheet']}]"
        labels = merged_headers(worksheet, spec["header_rows"], worksheet.max_column)
    finally:
        workbook.close()

    # The header signature check. A re-export that moves a column fails here, naming it.
    for col, expected in spec["expect"].items():
        found = labels[col] if col < len(labels) else "(no such column)"
        if found != expected:
            problems.append(
                f"{where}: column {col} header is {found!r}, expected {expected!r} — the source "
                "has been reshaped; re-read it before trusting this extract"
            )

    # The title block is not always in column 0 — TB puts it in column 3 or 4 and Demographics
    # in column 1 — so take the first non-empty cell of each title row rather than a fixed column.
    title = " | ".join(
        filter(None, (next((text(c) for c in rows[i] if text(c)), "") for i in (1, 2)))
    )

    resolver.reset()
    out_rows: list[dict] = []
    over_100 = Counter()
    not_reported = Counter()
    unresolved: list[str] = []
    data_rows = 0
    by_level = Counter()

    for row in rows[spec["data_start"]:]:
        if not is_data_row(row):
            continue
        source_psgc = raw_psgc(row[spec["psgc_col"]]) if len(row) > spec["psgc_col"] else ""
        area_name = text(row[spec["name_col"]]) if len(row) > spec["name_col"] else ""
        if looks_like_note(source_psgc, area_name):
            continue

        data_rows += 1

        if not source_psgc and area_name.upper() in {"PHILIPPINES", "PHILIPPINES "}:
            geo_code, geo_level, how = "PH", "national", "national"
            resolver.stats["national"] += 1
        else:
            resolved = resolver.resolve(source_psgc, area_name)
            if not resolved:
                unresolved.append(f"{source_psgc or '(blank)'} {area_name!r}")
                continue
            geo_code, geo_level, how = resolved

        by_level[geo_level] += 1

        def cell(col):
            return number(row[col], problems, f"{where} {area_name} col {col}") if col < len(
                row
            ) else None

        if spec["kind"] == "workforce":
            population = as_int(cell(spec["population_col"]))
            households = as_int(cell(spec["households_col"]))
            for i, (source_label, cadre) in enumerate(CADRES.items()):
                base = spec["cadre_first_col"] + i * spec["cadre_stride"]
                if labels[base] != f"{source_label} || LGU Hired":
                    problems.append(
                        f"{where}: cadre block {i} at column {base} is {labels[base]!r}, expected "
                        f"{source_label!r} — the cadre order has changed"
                    )
                out_rows.append(
                    {
                        "cadre": cadre,
                        "geo_code": geo_code,
                        "geo_level": geo_level,
                        "lgu_hired": as_int(cell(base)),
                        "doh_hired": as_int(cell(base + 1)),
                        "total": as_int(cell(base + 2)),
                        # base + 3 is the source's Ratio column: population divided by total.
                        # Not stored — it is recomputable from two columns that are, and
                        # Decision 4's rule is to store the counts.
                        "population_2025": population,
                        "households_2025": households,
                        "source_psgc": source_psgc,
                        "source_area_name": area_name,
                        "geo_resolution": how,
                    }
                )
            continue

        for (indicator_key, breakdown), cols in spec["measures"].items():
            numerator = as_int(cell(cols["numerator"])) if "numerator" in cols else None
            denominator = as_int(cell(cols["denominator"])) if "denominator" in cols else None
            rate = cell(cols["rate_pct"]) if "rate_pct" in cols else None
            if numerator is None and denominator is None and rate is None:
                # Nothing published for this area/indicator/breakdown. Decision 7: a missing row
                # is "not reported", which is not the same as a zero, so no row is written and
                # nothing is invented.
                not_reported[indicator_key] += 1
                continue
            flag = is_over_100(indicator_key, rate)
            if flag:
                over_100[indicator_key] += 1
            out_rows.append(
                {
                    "indicator_key": indicator_key,
                    "breakdown": breakdown,
                    "geo_code": geo_code,
                    "geo_level": geo_level,
                    "numerator": numerator,
                    "denominator": denominator,
                    "rate_pct": rate,
                    "over_100": flag,
                    "source_psgc": source_psgc,
                    "source_area_name": area_name,
                    "geo_resolution": how,
                }
            )

    if unresolved:
        problems.append(
            f"{where}: {len(unresolved)} rows whose geography could not be resolved by any of "
            f"the three stages, e.g. {unresolved[:5]}"
        )

    declared = set(spec["expect"]) | {spec["psgc_col"], spec["name_col"]}
    if spec["kind"] == "workforce":
        declared |= {spec["population_col"], spec["households_col"]}
        for i in range(len(CADRES)):
            base = spec["cadre_first_col"] + i * spec["cadre_stride"]
            declared |= {base, base + 1, base + 2}
    else:
        for cols in spec["measures"].values():
            declared |= set(cols.values())
    dropped = {
        col: spec.get("dropped", {}).get(col, labels[col])
        for col in range(len(labels))
        if col not in declared and labels[col]
    }

    return {
        "spec": spec,
        "where": where,
        "title": title,
        "rows": out_rows,
        "data_rows": data_rows,
        "by_level": by_level,
        "over_100": over_100,
        "not_reported": not_reported,
        "dropped": dropped,
        "header_signature": {c: labels[c] for c in sorted(spec["expect"])},
        "geo_stats": dict(resolver.stats),
        "geo_repairs": list(resolver.repairs),
        "geo_overrides": list(resolver.overrides),
        "geo_naming_differences": set(resolver.naming_differences),
    }


# --------------------------------------------------------------------------- #
# Reconciliation (Decision 4: publish the residual, never pick a side)          #
# --------------------------------------------------------------------------- #
def reconcile(indicator_rows: list[dict]) -> list[dict]:
    """Per indicator and breakdown: Σ children vs the published parent, at each step up the tree.

    Rollups are never recomputed from leaves for display, because the source publishes its own
    subtotals and its leaf table is known to be incomplete — nationally the citymun rows fall
    about 6,400 short of the published national figure in Demographics. So this measures the gap
    instead of choosing a side, and the residual goes to the report and to
    `ref_fhsis_reconciliation` as a stated fact.
    """
    published: dict[tuple, int] = {}
    for row in indicator_rows:
        if row["numerator"] is None:
            continue
        published[
            (row["indicator_key"], row["breakdown"], row["geo_level"], row["geo_code"])
        ] = row["numerator"]

    def parent_of(level: str, code: str):
        if level == "citymun":
            return "province", code[:5]
        if level == "province":
            return "region", code[:2]
        if level == "region":
            return "national", "PH"
        return None, None

    sums: dict[tuple, int] = defaultdict(int)
    counts: dict[tuple, int] = defaultdict(int)
    for (indicator, breakdown, level, code), value in published.items():
        parent_level, parent_code = parent_of(level, code)
        if not parent_level:
            continue
        sums[(indicator, breakdown, parent_level, parent_code)] += value
        counts[(indicator, breakdown, parent_level, parent_code)] += 1

    out = []
    for key, child_sum in sorted(sums.items()):
        indicator, breakdown, parent_level, parent_code = key
        parent_value = published.get(key)
        if parent_value is None:
            continue
        out.append(
            {
                "indicator_key": indicator,
                "breakdown": breakdown,
                "parent_level": parent_level,
                "parent_geo_code": parent_code,
                "child_count": counts[key],
                "child_sum": child_sum,
                "published_parent": parent_value,
                "residual": parent_value - child_sum,
            }
        )
    return out


# --------------------------------------------------------------------------- #
# Output                                                                       #
# --------------------------------------------------------------------------- #
def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    c: (
                        ""
                        if row.get(c) is None
                        else ("true" if row[c] is True else "false" if row[c] is False else row[c])
                    )
                    for c in columns
                }
            )


def national_check(indicator_rows: list[dict], workforce_rows: list[dict]) -> list[str]:
    """The figures the plan names as load-blocking, checked before anything is written."""
    problems = []
    index = {
        (r["indicator_key"], r["breakdown"]): r
        for r in indicator_rows
        if r["geo_level"] == "national"
    }
    for (indicator, breakdown, field), expected in NATIONAL_CHECKS.items():
        row = index.get((indicator, breakdown))
        found = row[field] if row else None
        if found != expected:
            problems.append(
                f"national {indicator}/{breakdown} {field} is {found!r}, expected {expected:,} — "
                "the source no longer reproduces the figure this load is pinned to"
            )
    national_workforce = [r for r in workforce_rows if r["geo_level"] == "national"]
    if national_workforce:
        population = national_workforce[0]["population_2025"]
        households = national_workforce[0]["households_2025"]
        if population != NATIONAL_POPULATION_2025:
            problems.append(
                f"national Population 2025 is {population!r}, expected "
                f"{NATIONAL_POPULATION_2025:,}"
            )
        if households != NATIONAL_HOUSEHOLDS_2025:
            problems.append(
                f"national Number of Household Estimates is {households!r}, expected "
                f"{NATIONAL_HOUSEHOLDS_2025:,}"
            )
    return problems


def bhw_check(indicator_rows: list[dict], workforce_rows: list[dict]) -> list[str]:
    """Decision 2, enforced on the extract rather than trusted.

    The cleaner is the only place the BHW columns are ever visible, so this is where their
    absence is proved: no cadre may be a BHW, and no indicator key may name one. The loader
    asserts the same thing against the database and the table carries a CHECK constraint; three
    independent guards, because the rule is an owner decision and not a preference.
    """
    problems = []
    bad_cadres = sorted({r["cadre"] for r in workforce_rows if "bhw" in r["cadre"].lower()})
    if bad_cadres:
        problems.append(f"BHW cadre rows reached the extract: {bad_cadres}")
    bad_indicators = sorted(
        {r["indicator_key"] for r in indicator_rows if "bhw" in r["indicator_key"].lower()}
    )
    if bad_indicators:
        problems.append(f"BHW indicator rows reached the extract: {bad_indicators}")
    return problems


def render_report(results: list[dict], manifest: dict, reconciliation: list[dict],
                  outputs: dict) -> str:
    lines: list[str] = []
    add = lines.append

    add("# FHSIS 2025 — cleaning report")
    add("")
    add(
        "Generated by `python ingestion/clean_fhsis.py`. Every figure below is read from the "
        "workbooks named in the provenance table, not from a catalogue. Plan: "
        "`docs/FHSIS_2025_PLAN.md` (increment F1)."
    )
    add("")
    add(
        "This report is the diff target for a future pull. The archive is mutable — DOH keeps "
        "adding to and re-touching the 2025 folder — so a re-run that changes any figure here is "
        "a new version of the dataset, not a correction of this one, and bumps "
        "`dim_dataset.last_updated_at` through `bumpDatasetVersion` (plan Decision 9)."
    )
    add("")

    add("## Provenance, per workbook")
    add("")
    add(
        "| Workbook | Drive file id | Modified | Sheet read | Data rows | Period |"
    )
    add("|---|---|---|---|---|---|")
    by_key = {w["key"]: w for w in manifest["workbooks"]}
    for result in results:
        spec = result["spec"]
        entry = by_key.get(spec["workbook"], {})
        modified = entry.get("drive_modified_at") or "(not stated)"
        if entry.get("drive_modified_precision") == "day":
            modified += " (day precision — a native Google Sheet, see fetch_fhsis.py)"
        add(
            f"| `{spec['file']}` | `{entry.get('file_id', '?')}` | {modified} | "
            f"`{spec['sheet']}` | {result['data_rows']:,} | {spec['period']} |"
        )
    add("")
    add(f"Archive: `{manifest['archive_shortlink']}` → folder `{manifest['archive_folder']}`, "
        f"2025 Annual Excel folder `{manifest['annual_excel_2025_folder']}`. "
        f"Retrieved {manifest['retrieved_at']}.")
    add("")

    add("## Workbooks and sheets deliberately not loaded")
    add("")
    add(
        "Each is a property of the source found by reading it, and each corrects something the "
        "plan assumed before the files were opened. The plan has been amended to match."
    )
    add("")
    for key, why in EXCLUDED_WORKBOOKS.items():
        entry = by_key.get(key, {})
        add(f"- **`{entry.get('filename', key)}`** — {why}")
    for key, why in EXCLUDED_SHEETS.items():
        add(f"- **{key}** — {why}")
    add("")

    add("## Header signatures")
    add("")
    add(
        "The merged header rows this extract was written against, per sheet, for the columns it "
        "reads. A re-export that reshapes a table fails the run naming the column rather than "
        "reading the wrong one."
    )
    add("")
    for result in results:
        spec = result["spec"]
        add(f"- `{spec['file']}` [`{spec['sheet']}`] — header rows "
            f"{'+'.join(str(r + 1) for r in spec['header_rows'])}, data from row "
            f"{spec['data_start'] + 1}")
        add(f"  - title: {result['title']}")
        for col, label in result["header_signature"].items():
            add(f"  - column {col}: `{label}`")
    add("")

    add("## Geography resolution")
    add("")
    add(
        "Three stages, counted separately so the mix is visible. Stage 2 and 3 exist because "
        "about 70 rows per sheet carry a PSGC with its leading zeros stripped and the value "
        "right-padded back to width. Every loaded sheet ends with **0 unresolved**."
    )
    add("")
    add("| Sheet | national | direct | shift repair | name (parent-scoped) | unresolved |")
    add("|---|---|---|---|---|---|")
    for result in results:
        stats = result["geo_stats"]
        add(
            f"| `{result['spec']['file'][:34]}` [`{result['spec']['sheet']}`] | "
            f"{stats.get('national', 0)} | {stats.get('direct', 0):,} | "
            f"{stats.get('shift_repair', 0)} | {stats.get('name_scoped', 0)} | "
            f"{stats.get('unresolved', 0)} |"
        )
    add("")
    repairs = sorted({r for result in results for r in result["geo_repairs"]})
    add(f"### The {len(repairs)} distinct repairs, in full")
    add("")
    add(
        "Listed rather than counted, because a repaired code is a claim about which place a "
        "figure belongs to and a reader must be able to check every one."
    )
    add("")
    for line in repairs:
        add(f"- {line}")
    add("")

    add("## Rows written")
    add("")
    add("| Output | Rows |")
    add("|---|---|")
    for path, count in outputs.items():
        add(f"| `{path}` | {count:,} |")
    add("")
    add("### By geographic level")
    add("")
    add("| Sheet | national | region | province | citymun |")
    add("|---|---|---|---|---|")
    for result in results:
        by_level = result["by_level"]
        add(
            f"| `{result['spec']['file'][:34]}` [`{result['spec']['sheet']}`] | "
            f"{by_level.get('national', 0)} | {by_level.get('region', 0)} | "
            f"{by_level.get('province', 0)} | {by_level.get('citymun', 0):,} |"
        )
    add("")

    add("## Coverage above 100%")
    add("")
    add(
        "Stored **as published**, never capped, with `over_100 = true`. The denominator is "
        "present, so the overshoot is explainable — more services recorded than the projected "
        "eligible population — and every surface renders it with the † and footnote "
        "`docs/UUC_PHC_2025_PLAN.md` U3 established. The flag is set only for percentage "
        "indicators: the TB notification rates are per 100,000 and are normally in the hundreds, "
        "so marking them would put a warning on an ordinary figure."
    )
    add("")
    add("| Indicator | Unit | Rows over 100 |")
    add("|---|---|---|")
    totals = Counter()
    for result in results:
        totals.update(result["over_100"])
    for key, meta in INDICATORS.items():
        unit = "percent" if meta["unit"] == UNIT_PERCENT else "per 100,000"
        add(f"| `{key}` | {unit} | {totals.get(key, 0):,} |")
    add("")

    add("## Not reported, which is not zero")
    add("")
    add(
        "An (area, indicator, breakdown) with no published figure gets **no row**. Decision 7: a "
        "city with a row and a zero delivered zero and the page says so; a city with no row is "
        "not reporting and the page says that instead. The two are never conflated, which is why "
        "this extract does not `left join dim_geo` zeros into existence the way "
        "`agg_nhfr_counts` does."
    )
    add("")
    add("| Indicator | (area, breakdown) pairs with nothing published |")
    add("|---|---|")
    missing = Counter()
    for result in results:
        missing.update(result["not_reported"])
    for key in INDICATORS:
        add(f"| `{key}` | {missing.get(key, 0):,} |")
    add("")

    add("## Columns read, and columns dropped")
    add("")
    add(
        "Only declared columns are carried. Everything else is dropped and named here — the "
        "opposite of dropping the columns labelled `DQC`, because `DQC` is not a reliable label: "
        "8ANC uses it, 4ANC leaves the same check block unlabelled with sub-headers like "
        "`C>=H`, and FIC/CIC has two unlabelled trailing columns. Naming every dropped column is "
        "what stops an unnoticed indicator hiding among them."
    )
    add("")
    for result in results:
        spec = result["spec"]
        add(f"### `{spec['file']}` [`{spec['sheet']}`]")
        add("")
        if not result["dropped"]:
            add("No columns dropped.")
            add("")
            continue
        for col, why in sorted(result["dropped"].items()):
            add(f"- column {col}: {why}")
        add("")

    add("## Subtotal reconciliation")
    add("")
    add(
        "Σ children vs the published parent, per indicator and breakdown, at each step up the "
        "tree. The source publishes its own subtotals and its leaf table is incomplete, so "
        "rollups are never recomputed from leaves for display — this measures the gap and "
        "publishes it instead of choosing a side. `ref_fhsis_reconciliation` exposes the same "
        "numbers to the running site."
    )
    add("")
    national = [r for r in reconciliation if r["parent_level"] == "national"]
    add("### Region sums vs the published national row")
    add("")
    add("| Indicator | Breakdown | Regions | Σ regions | Published national | Residual |")
    add("|---|---|---|---|---|---|")
    for row in sorted(national, key=lambda r: (r["indicator_key"], r["breakdown"])):
        add(
            f"| `{row['indicator_key']}` | {row['breakdown']} | {row['child_count']} | "
            f"{row['child_sum']:,} | {row['published_parent']:,} | {row['residual']:,} |"
        )
    add("")
    add("### City/municipality sums vs their published province rows")
    add("")
    add(
        "Per indicator: how many provinces reconcile exactly, and every one that does not. A "
        "province-tier node with no children of its own — the 33 highly urbanised cities that "
        "`dim_geo` places at province level, City of Makati and the rest — is a leaf, not a "
        "parent, and is not compared with anything."
    )
    add("")
    add("| Indicator | Breakdown | Provinces compared | Reconcile exactly | Residual rows |")
    add("|---|---|---|---|---|")
    citymun_gap = [r for r in reconciliation if r["parent_level"] == "province"]
    by_indicator: dict[tuple, list] = defaultdict(list)
    for row in citymun_gap:
        by_indicator[(row["indicator_key"], row["breakdown"])].append(row)
    for (indicator, breakdown), rows in sorted(by_indicator.items()):
        exact = sum(1 for r in rows if r["residual"] == 0)
        add(
            f"| `{indicator}` | {breakdown} | {len(rows)} | {exact} | {len(rows) - exact} |"
        )
    add("")
    add("#### Every province whose cities do not sum to its published total")
    add("")
    add(
        "In every case checked, the gap is a city that `dim_geo` nests under the province and "
        "the source's province row excludes — an independent component or highly urbanised city "
        "that reports separately. For FIC the five are City of Cotabato, City of Dagupan, City "
        "of Naga, City of Santiago and Ormoc City, and each province's residual equals that "
        "city's own figure exactly. This is the reason Decision 4 forbids deriving a provincial "
        "figure by summing its cities: the two are answers to different questions, and the "
        "source publishes the one it means."
    )
    add("")
    add("| Indicator | Breakdown | Province | Children | Σ children | Published | Residual |")
    add("|---|---|---|---|---|---|---|")
    off = [r for r in citymun_gap if r["residual"] != 0]
    for row in sorted(
        off, key=lambda r: (r["indicator_key"], r["breakdown"], r["parent_geo_code"])
    ):
        add(
            f"| `{row['indicator_key']}` | {row['breakdown']} | `{row['parent_geo_code']}` | "
            f"{row['child_count']} | {row['child_sum']:,} | {row['published_parent']:,} | "
            f"{row['residual']:,} |"
        )
    add("")

    add("## Where the printed name disagreed with the printed code")
    add("")
    add(
        "The area name is used as a **check on** the PSGC, never as the join key. It is what "
        "catches the one block in this release whose codes are swapped: the province row "
        "\"Surigao del Norte\" carries `1606701000`, which is Alegria's code, and the row "
        "\"Alegria\" carries the province's. Honouring the code alone would file 3,118 antenatal "
        "visits under a municipality that had 9, and nothing would look wrong — both codes "
        "resolve cleanly."
    )
    add("")
    overrides = sorted({o for result in results for o in result["geo_overrides"]})
    add(f"### {len(overrides)} rows where the name won")
    add("")
    add(
        "The override fires only when the printed name resolves, inside the same parent, to a "
        "*different* geography."
    )
    add("")
    for line in overrides:
        add(f"- {line}")
    add("")
    differences = sorted({d for result in results for d in result["geo_naming_differences"]})
    add(f"### {len(differences)} rows where the wording differs but the code stands")
    add("")
    add(
        "Label style, spelling and renaming — not coding errors. The name resolves to nothing "
        "else, so the code is honoured and the row is recorded here rather than changed."
    )
    add("")
    for line in differences:
        add(f"- {line}")
    add("")

    add("## The BHW rule, and where it is enforced")
    add("")
    add(
        "FHSIS never supplies a BHW count (plan Decision 2). The Demographics `Health Workers` "
        "sheet publishes `Active Barangay Health Workers (BHW)` and its ratio; both are dropped "
        "here, so they exist in no CSV and therefore in no table. The source's own numbers are "
        "the argument: 270,766 nationally, 4,454 in NCR and **1 in Las Piñas** — a tally of what "
        "LGUs filed through their RHUs, not a census. This site's BHW figure is `bhw-2025` with "
        "the StepZero quick-count as the universe."
    )
    add("")
    add("Enforced three times, independently:")
    add("")
    add("1. `clean_fhsis.py` drops the columns and asserts no BHW cadre or indicator key reaches "
        "the extract.")
    add("2. `fact_fhsis_workforce` carries `check (cadre <> 'bhw')`.")
    add("3. `ingest_fhsis.py` asserts zero BHW cadre rows before it commits.")
    add("")
    add(
        "The registry note says the cadre *does not exist in this table, not merely hidden*, so "
        "the assistant never answers as though something were being withheld."
    )
    add("")

    add("## Denominators")
    add("")
    add(
        "Every denominator here is the **source's own** (plan Decision 5). FHSIS's "
        "`Population 2025` and `Number of Household Estimates` are DOH projections carried for "
        "DOH's own ratio arithmetic; they are loaded so a published rate stays recomputable "
        "against the base it was computed on, and they are this dataset's base and nothing "
        "else's. No per-capita or per-household figure elsewhere on the site moves onto them: "
        "the site's per-capita denominator is StepZero's own self-reported population, with "
        "`agg_population` as the fallback, per the owner decision of 2026-09-06 in "
        "`docs/DECISIONS.md`."
    )
    add("")
    add(
        "Two indicators are the exception worth stating: `tb_tsr_dstb` and `tb_tsr_mdrtb` divide "
        "by a **case count**, not a population — the number of TB cases registered — because a "
        "treatment success rate is a proportion of patients treated, not of people living there."
    )
    add("")

    return "\n".join(lines) + "\n"


# --------------------------------------------------------------------------- #
def build_summary(results: list[dict], outputs: dict, reconciliation: list[dict],
                  manifest: dict) -> dict:
    """The counters the loader re-checks, from the same run that wrote the CSVs."""
    over_100: Counter = Counter()
    not_reported: Counter = Counter()
    for result in results:
        over_100.update(result["over_100"])
        not_reported.update(result["not_reported"])
    return {
        "generated_by": "ingestion/clean_fhsis.py",
        "source_retrieved_at": manifest["retrieved_at"],
        "workbooks": {
            entry["key"]: {
                "file_id": entry["file_id"],
                "drive_modified_at": entry["drive_modified_at"],
                "content_digest": entry["content_digest"],
            }
            for entry in manifest["workbooks"]
        },
        "output_rows": {Path(path).name: count for path, count in outputs.items()},
        "over_100_by_indicator": {k: over_100.get(k, 0) for k in INDICATORS},
        "not_reported_by_indicator": {k: not_reported.get(k, 0) for k in INDICATORS},
        "sheet_rows": {
            f"{r['spec']['file']} [{r['spec']['sheet']}]": {
                "source_rows": r["data_rows"],
                "long_rows": len(r["rows"]),
                "geo": dict(r["geo_stats"]),
                "by_level": dict(r["by_level"]),
            }
            for r in results
        },
        "reconciliation": {
            "region_vs_national_nonzero": sum(
                1 for r in reconciliation if r["parent_level"] == "national" and r["residual"]
            ),
            "province_rows": sum(1 for r in reconciliation if r["parent_level"] == "province"),
            "province_rows_reconciling": sum(
                1
                for r in reconciliation
                if r["parent_level"] == "province" and r["residual"] == 0
            ),
        },
    }


def clean(src_dir: Path, out_dir: Path, dim_geo: dict, report_path: Path,
          summary_path: Path) -> dict:
    manifest_path = src_dir / MANIFEST_NAME
    if not manifest_path.exists():
        raise SystemExit(
            f"{manifest_path} not found — run `python ingestion/fetch_fhsis.py` first. The "
            "manifest is this dataset's provenance record, so the report cannot be written "
            "without it."
        )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    resolver = GeoResolver(dim_geo)
    problems: list[str] = []
    results = [extract_sheet(spec, src_dir, resolver, problems) for spec in SHEETS]

    indicator_rows: list[dict] = []
    workforce_rows: list[dict] = []
    for result in results:
        if result["spec"]["kind"] == "workforce":
            workforce_rows.extend(result["rows"])
        else:
            indicator_rows.extend(result["rows"])

    problems.extend(national_check(indicator_rows, workforce_rows))
    problems.extend(bhw_check(indicator_rows, workforce_rows))

    seen = set()
    for row in indicator_rows:
        key = (row["indicator_key"], row["breakdown"], row["geo_code"])
        if key in seen:
            problems.append(
                f"duplicate (indicator, breakdown, geo_code): {key} — two source rows resolved "
                "to the same place, which the fact table's unique constraint would reject"
            )
        seen.add(key)
    seen_workforce = set()
    for row in workforce_rows:
        key = (row["cadre"], row["geo_code"])
        if key in seen_workforce:
            problems.append(f"duplicate (cadre, geo_code): {key}")
        seen_workforce.add(key)

    unknown = {r["indicator_key"] for r in indicator_rows} - set(INDICATORS)
    if unknown:
        problems.append(f"indicator keys not in the dictionary: {sorted(unknown)}")
    unknown_breakdowns = {r["breakdown"] for r in indicator_rows} - set(BREAKDOWNS)
    if unknown_breakdowns:
        problems.append(f"breakdowns outside the vocabulary: {sorted(unknown_breakdowns)}")

    if problems:
        shown = "\n".join(f"  - {p}" for p in problems[:50])
        more = f"\n  ... and {len(problems) - 50} more" if len(problems) > 50 else ""
        raise SystemExit(
            "Cleaning aborted — the source disagrees with what this script was written for:\n"
            + shown
            + more
        )

    outputs: dict[str, int] = {}
    out_dir.mkdir(parents=True, exist_ok=True)
    for area in sorted({r["spec"]["area"] for r in results if r["spec"]["kind"] == "indicator"}):
        rows = [
            row
            for result in results
            if result["spec"]["kind"] == "indicator" and result["spec"]["area"] == area
            for row in result["rows"]
        ]
        path = out_dir / f"fhsis_2025_{area}_cleaned.csv"
        write_csv(path, OUT_COLUMNS_INDICATOR, rows)
        outputs[str(path)] = len(rows)

    workforce_path = out_dir / "fhsis_2025_workforce_cleaned.csv"
    write_csv(workforce_path, OUT_COLUMNS_WORKFORCE, workforce_rows)
    outputs[str(workforce_path)] = len(workforce_rows)

    reconciliation = reconcile(indicator_rows)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        render_report(results, manifest, reconciliation, outputs), encoding="utf-8"
    )
    outputs[str(report_path)] = len(reconciliation)

    summary_path.write_text(
        json.dumps(build_summary(results, outputs, reconciliation, manifest), indent=2) + "\n",
        encoding="utf-8",
    )

    return {
        "results": results,
        "indicator_rows": len(indicator_rows),
        "workforce_rows": len(workforce_rows),
        "outputs": outputs,
        "reconciliation": reconciliation,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--src-dir", default=SRC_DIR_DEFAULT, help="Where fetch_fhsis.py wrote")
    parser.add_argument("--out-dir", default=OUT_DIR_DEFAULT, help="Where to write cleaned CSVs")
    parser.add_argument("--report", default=REPORT_DEFAULT, help="Where to write the report")
    parser.add_argument(
        "--summary",
        default=SUMMARY_DEFAULT,
        help="Where to write the machine-readable summary the loader re-checks",
    )
    parser.add_argument("--dim-geo-csv", help="dim_geo export CSV (geo_code, geo_level, geo_name)")
    parser.add_argument("--database-url", help="Read dim_geo live from this Postgres URL instead")
    args = parser.parse_args()

    if not args.dim_geo_csv and not args.database_url:
        parser.error("pass --dim-geo-csv (offline) or --database-url")

    dim_geo = (
        load_dim_geo_from_db(args.database_url)
        if args.database_url
        else load_dim_geo_csv(args.dim_geo_csv)
    )
    print(f"dim_geo: {len(dim_geo):,} geographies")

    stats = clean(
        Path(args.src_dir), Path(args.out_dir), dim_geo, Path(args.report), Path(args.summary)
    )

    print(f"\nWrote {stats['indicator_rows']:,} indicator rows and "
          f"{stats['workforce_rows']:,} workforce rows")
    for path, count in stats["outputs"].items():
        print(f"  {path}: {count:,}")
    print()
    for result in stats["results"]:
        spec = result["spec"]
        print(f"  {spec['file'][:34]:<36} [{spec['sheet']:<14}] "
              f"{result['data_rows']:>5,} source rows -> {len(result['rows']):>6,} long rows  "
              f"{dict(result['geo_stats'])}")


if __name__ == "__main__":
    main()
